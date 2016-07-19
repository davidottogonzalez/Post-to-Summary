from openpyxl import load_workbook
from openpyxl.styles import Font
import shutil, os, glob, collections, re

source_wb = None
summary_wb = None


def format_cell(cell, header):
    if 'pct' in header:
        cell.number_format = '##0.00%'
    elif 'dollars' in header:
        cell.number_format = '$#,##0.00'
    elif 'time' in header or 'date' in header:
        cell.number_format = 'MM-DD-YYYY HH:MM:SS'
    else:
        cell.number_format = '#,##0.00'


def setup(source_filename, equiv):
    if equiv:
        new_template = source_filename.replace(".xlsx", "_summary_equiv.xlsx")
    else:
        new_template = source_filename.replace(".xlsx", "_summary_unequiv.xlsx")
    shutil.copy('template.xlsx', new_template)

    global source_wb
    global summary_wb

    source_wb = load_workbook('./preprocessed/' + source_filename)
    summary_wb = load_workbook(new_template)

    return new_template


def process_summary_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_sheet = source_wb.get_sheet_by_name('Summary')
    source_spot_sheet = source_wb.get_sheet_by_name('Spot Detail')
    summary_sheet = summary_wb.get_sheet_by_name('Summary Metrics')
    program_metrics_sheet = summary_wb.get_sheet_by_name('Program Metrics')
    start_row = 0
    net_sum_headers_row = 0
    network_list = []

    summary_sheet['A1'] = 'EQUIVALIZED - Campaign Summary' if equiv else 'UNEQUIVALIZED - Campaign Summary'
    summary_sheet.cell(row=summary_sheet.max_row + 2, column=1).value = 'Network Summary'

    # Get network order
    for row_num in range(3, source_sheet.max_row + 1):
        network_list.append((source_sheet['A'+str(row_num)].value,row_num))
    network_list = sorted(network_list, key=lambda tup: tup[0])
    row_list = [1,2]

    for network in network_list:
        row_list.append(network[1])

    for row_num in row_list:
        if start_row == 0:
            start_row = summary_sheet.max_row + 2
            net_sum_headers_row = summary_sheet.max_row + 2
        else:
            start_row = summary_sheet.max_row + 1

        write_col = 1
        for col_num in range(1, source_sheet.max_column + 1):
            if source_sheet.cell(row=1, column=col_num).value in ['num_spots', 'equiv_units', 'total_impressions',
                                                                  'total_unequiv_impressions',
                                                                  'total_unequiv_frequency', 'GRPs', 'GRPs_unequiv',
                                                                  'target_impressions', 'target_unequiv_impressions',
                                                                  'target_unequiv_frequency', 'TRPs', 'TRPs_unequiv',
                                                                  'target_index_impressions',
                                                                  'target_index_unequiv_impressions', 'tCPM']:
                if equiv:
                    if source_sheet.cell(row=1, column=col_num).value in ['equiv_units', 'total_impressions', 'GRPs',
                                                                          'target_impressions', 'TRPs',
                                                                          'target_index_impressions', 'tCPM']:
                        summary_sheet.cell(row=start_row, column=write_col).value = source_sheet.cell(row=row_num,
                                                                                                      column=col_num).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_sheet.cell(row=1, column=col_num).value == 'total_unequiv_frequency':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'total_frequency'
                        else:
                            summary_sheet.cell(row=start_row, column=write_col).value = \
                                source_sheet.cell(row=row_num, column=5).value / source_sheet.cell(row=row_num,
                                                                                                   column=7).value

                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_sheet.cell(row=1, column=col_num).value == 'target_unequiv_frequency':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'target_frequency'
                        else:
                            summary_sheet.cell(row=start_row, column=write_col).value = \
                                source_sheet.cell(row=row_num, column=17).value / source_sheet.cell(row=row_num,
                                                                                                    column=19).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                else:
                    if source_sheet.cell(row=1, column=col_num).value in ['num_spots', 'total_unequiv_impressions',
                                                                          'GRPs_unequiv', 'target_unequiv_impressions',
                                                                          'TRPs_unequiv',
                                                                          'target_index_unequiv_impressions',
                                                                          'total_unequiv_frequency',
                                                                          'target_unequiv_frequency']:
                        summary_sheet.cell(row=start_row, column=write_col).value = source_sheet.cell(row=row_num,
                                                                                                      column=col_num).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_sheet.cell(row=1, column=col_num).value == 'tCPM':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'tCPM_unequiv'
                        else:
                            summary_sheet.cell(row=start_row, column=write_col).value = \
                                source_sheet.cell(row=row_num, column=4).value * 1000 / source_sheet.cell(row=row_num,
                                                                                                          column=18).value
                            format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                        source_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
            else:
                summary_sheet.cell(row=start_row, column=write_col).value = source_sheet.cell(row=row_num,
                                                                                              column=col_num).value

                format_cell(summary_sheet.cell(row=start_row, column=write_col),
                            source_sheet.cell(row=1, column=col_num).value)
                write_col += 1

    start_row = 0
    p_start_row = 3

    summary_sheet.cell(row=summary_sheet.max_row + 2, column=1).value = 'Spot Detail'
    program_metrics_sheet['A1'] = 'EQUIVALIZED - Program Metrics' if equiv else 'UNEQUIVALIZED - Program Metrics'

    for row_num in range(1, source_spot_sheet.max_row + 1):
        if start_row == 0:
            start_row = summary_sheet.max_row + 2
        else:
            start_row = summary_sheet.max_row + 1

        write_col = 1
        for col_num in range(1, source_spot_sheet.max_column + 1):
            if source_spot_sheet.cell(row=1, column=col_num).value in ['num_spots', 'equiv_units', 'total_impressions',
                                                                       'total_unequiv_impressions',
                                                                       'total_unequiv_frequency', 'GRPs',
                                                                       'GRPs_unequiv',
                                                                       'target_impressions',
                                                                       'target_unequiv_impressions',
                                                                       'target_unequiv_frequency', 'TRPs',
                                                                       'TRPs_unequiv',
                                                                       'target_index_impressions',
                                                                       'target_index_unequiv_impressions', 'tCPM']:
                if equiv:
                    if source_spot_sheet.cell(row=1, column=col_num).value in ['equiv_units', 'total_impressions',
                                                                               'GRPs',
                                                                               'target_impressions', 'TRPs',
                                                                               'target_index_impressions', 'tCPM']:
                        summary_sheet.cell(row=start_row, column=write_col).value = source_spot_sheet.cell(row=row_num,
                                                                                                           column=col_num).value
                        program_metrics_sheet.cell(row=p_start_row, column=write_col).value = source_spot_sheet.cell(
                            row=row_num,
                            column=col_num).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_spot_sheet.cell(row=1, column=col_num).value == 'total_unequiv_frequency':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'total_frequency'
                            program_metrics_sheet.cell(row=p_start_row, column=write_col).value = 'total_frequency'
                        else:
                            summary_sheet.cell(row=start_row, column=write_col).value = \
                                source_spot_sheet.cell(row=row_num, column=34).value / source_spot_sheet.cell(
                                    row=row_num,
                                    column=36).value
                            program_metrics_sheet.cell(row=p_start_row, column=write_col).value = \
                                source_spot_sheet.cell(row=row_num, column=34).value / source_spot_sheet.cell(
                                    row=row_num,
                                    column=36).value

                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_spot_sheet.cell(row=1, column=col_num).value == 'target_unequiv_frequency':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'target_frequency'
                            program_metrics_sheet.cell(row=p_start_row, column=write_col).value = 'target_frequency'
                        else:
                            summary_sheet.cell(row=start_row, column=write_col).value = \
                                source_spot_sheet.cell(row=row_num, column=38).value / source_spot_sheet.cell(
                                    row=row_num,
                                    column=40).value
                            program_metrics_sheet.cell(row=p_start_row, column=write_col).value = \
                                source_spot_sheet.cell(row=row_num, column=38).value / source_spot_sheet.cell(
                                    row=row_num,
                                    column=40).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)

                        write_col += 1
                else:
                    if source_spot_sheet.cell(row=1, column=col_num).value in ['num_spots', 'total_unequiv_impressions',
                                                                               'GRPs_unequiv',
                                                                               'target_unequiv_impressions',
                                                                               'TRPs_unequiv',
                                                                               'target_index_unequiv_impressions',
                                                                               'total_unequiv_frequency',
                                                                               'target_unequiv_frequency']:
                        summary_sheet.cell(row=start_row, column=write_col).value = source_spot_sheet.cell(row=row_num,
                                                                                                           column=col_num).value
                        program_metrics_sheet.cell(row=p_start_row, column=write_col).value = source_spot_sheet.cell(
                            row=row_num, column=col_num).value
                        format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                                    source_spot_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
                    if source_spot_sheet.cell(row=1, column=col_num).value == 'tCPM':
                        if row_num == 1:
                            summary_sheet.cell(row=start_row, column=write_col).value = 'tCPM_unequiv'
                            program_metrics_sheet.cell(row=p_start_row, column=write_col).value = 'tCPM_unequiv'
                        else:
                            if source_spot_sheet.cell(row=row_num, column=39).value != None:
                                summary_sheet.cell(row=start_row, column=write_col).value = \
                                    source_spot_sheet.cell(row=row_num,
                                                           column=26).value * 1000 / source_spot_sheet.cell(row=row_num,
                                                                                                            column=39).value
                                program_metrics_sheet.cell(row=p_start_row, column=write_col).value = \
                                    source_spot_sheet.cell(row=row_num,
                                                           column=26).value * 1000 / source_spot_sheet.cell(
                                        row=row_num, column=39).value
                                format_cell(summary_sheet.cell(row=start_row, column=write_col),
                                            source_spot_sheet.cell(row=1, column=col_num).value)
                                format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                                            source_spot_sheet.cell(row=1, column=col_num).value)
                        write_col += 1
            else:
                summary_sheet.cell(row=start_row, column=write_col).value = source_spot_sheet.cell(row=row_num,
                                                                                                   column=col_num).value
                program_metrics_sheet.cell(row=p_start_row, column=write_col).value = source_spot_sheet.cell(
                    row=row_num,
                    column=col_num).value

                format_cell(summary_sheet.cell(row=start_row, column=write_col),
                            source_spot_sheet.cell(row=1, column=col_num).value)
                format_cell(program_metrics_sheet.cell(row=p_start_row, column=write_col),
                            source_spot_sheet.cell(row=1, column=col_num).value)
                write_col += 1

        p_start_row += 1

    # start totals
    for row_num in range(1, source_sheet.max_row + 1):
        if source_sheet.cell(row=row_num, column=1).value == 'Total':
            for col_num in range(1, source_sheet.max_column + 1):
                if source_sheet.cell(row=1, column=col_num).value == 'total_impressions' and equiv:
                    summary_sheet.cell(row=6, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_unequiv_impressions' and not equiv:
                    summary_sheet.cell(row=6, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_reach':
                    summary_sheet.cell(row=8, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_reach_pct':
                    summary_sheet.cell(row=10, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_effective_reach':
                    summary_sheet.cell(row=12, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_effective_reach_pct':
                    summary_sheet.cell(row=13, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'total_reach_raw_count':
                    summary_sheet.cell(row=9, column=3).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_impressions' and equiv:
                    summary_sheet.cell(row=6, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_unequiv_impressions' and not equiv:
                    summary_sheet.cell(row=6, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_reach':
                    summary_sheet.cell(row=8, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_reach_pct':
                    summary_sheet.cell(row=10, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_effective_reach':
                    summary_sheet.cell(row=12, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_effective_reach_pct':
                    summary_sheet.cell(row=13, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_reach_raw_count':
                    summary_sheet.cell(row=9, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_index_impressions' and equiv:
                    summary_sheet.cell(row=7, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_index_unequiv_impressions' and not equiv:
                    summary_sheet.cell(row=7, column=2).value = source_sheet.cell(row=row_num, column=col_num).value
                if source_sheet.cell(row=1, column=col_num).value == 'target_index_reach':
                    summary_sheet.cell(row=11, column=2).value = source_sheet.cell(row=row_num, column=col_num).value

            # insert 1 for indexes of total
            summary_sheet.cell(row=7, column=3).value = 100.0
            summary_sheet.cell(row=11, column=3).value = 100.0
            summary_sheet.cell(row=14, column=3).value = 100.0

            # custom calculations
            # Index Effective Reach
            summary_sheet.cell(row=14, column=2).value = summary_sheet.cell(row=13,
                                                                            column=2).value / summary_sheet.cell(row=13,
                                                                                                                 column=3).value if summary_sheet.cell(
                row=13, column=2).value else 0
            # Average Frequency
            summary_sheet.cell(row=15, column=2).value = summary_sheet.cell(row=6, column=2).value / summary_sheet.cell(
                row=8, column=2).value
            summary_sheet.cell(row=15, column=3).value = summary_sheet.cell(row=6, column=3).value / summary_sheet.cell(
                row=8, column=3).value
            # Avg Freq % Diff v. Total
            summary_sheet.cell(row=16, column=2).value = (summary_sheet.cell(row=15,
                                                                             column=2).value - summary_sheet.cell(
                row=15, column=3).value) / summary_sheet.cell(row=15, column=3).value
            summary_sheet.cell(row=16, column=3).value = (summary_sheet.cell(row=15,
                                                                             column=3).value - summary_sheet.cell(
                row=15, column=3).value) / summary_sheet.cell(row=15, column=3).value

            # Projected Calculations
            summary_sheet.cell(row=4, column=2).value = summary_sheet.cell(row=8, column=2).value / summary_sheet.cell(
                row=10, column=2).value
            summary_sheet.cell(row=4, column=3).value = summary_sheet.cell(row=8, column=3).value / summary_sheet.cell(
                row=10, column=3).value
            summary_sheet.cell(row=5, column=2).value = summary_sheet.cell(row=4, column=2).value / summary_sheet.cell(
                row=4, column=3).value
            summary_sheet.cell(row=5, column=3).value = summary_sheet.cell(row=4, column=3).value / summary_sheet.cell(
                row=4, column=3).value

            # Format totals
            format_cell(summary_sheet.cell(row=4, column=2), '')
            format_cell(summary_sheet.cell(row=4, column=3), '')
            format_cell(summary_sheet.cell(row=6, column=2), '')
            format_cell(summary_sheet.cell(row=6, column=3), '')
            format_cell(summary_sheet.cell(row=7, column=2), '')
            format_cell(summary_sheet.cell(row=7, column=3), '')
            format_cell(summary_sheet.cell(row=8, column=2), '')
            format_cell(summary_sheet.cell(row=8, column=3), '')
            format_cell(summary_sheet.cell(row=9, column=2), '')
            format_cell(summary_sheet.cell(row=9, column=3), '')
            format_cell(summary_sheet.cell(row=11, column=2), '')
            format_cell(summary_sheet.cell(row=11, column=3), '')
            format_cell(summary_sheet.cell(row=12, column=2), '')
            format_cell(summary_sheet.cell(row=12, column=3), '')
            format_cell(summary_sheet.cell(row=14, column=2), '')
            format_cell(summary_sheet.cell(row=14, column=3), '')
            format_cell(summary_sheet.cell(row=15, column=2), '')
            format_cell(summary_sheet.cell(row=15, column=3), '')
            format_cell(summary_sheet.cell(row=5, column=2), 'pct')
            format_cell(summary_sheet.cell(row=5, column=3), 'pct')
            format_cell(summary_sheet.cell(row=10, column=2), 'pct')
            format_cell(summary_sheet.cell(row=10, column=3), 'pct')
            format_cell(summary_sheet.cell(row=13, column=2), 'pct')
            format_cell(summary_sheet.cell(row=13, column=3), 'pct')
            format_cell(summary_sheet.cell(row=16, column=2), 'pct')
            format_cell(summary_sheet.cell(row=16, column=3), 'pct')

            break

    summary_wb.save(filename)

    return True


def process_Network_Daypart_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_network_day_sheet = source_wb.get_sheet_by_name('Network Daypart')
    dest_network_day_sheet = summary_wb.get_sheet_by_name("Network Daypart")

    dest_network_day_sheet['A1'].font = Font(bold=True)
    dest_network_day_sheet['A1'] = 'EQUIVALIZED - Network Daypart Summary' if equiv else 'UNEQUIVALIZED - Network Daypart Summary'

    # Get first non-Total row
    first_non_total = 0
    for row_num in range(2, source_network_day_sheet.max_row + 1):
        if source_network_day_sheet['A'+str(row_num)].value != 'Total':
            first_non_total = row_num
            break

    # Get network order
    network_list = []
    for row_num in range(first_non_total, source_network_day_sheet.max_row + 1):
        network_list.append((source_network_day_sheet['A' + str(row_num)].value, row_num))
    network_list.sort()
    row_list = range(1,first_non_total)

    for network in network_list:
        row_list.append(network[1])

    write_row = 2
    for row_num in row_list:
        write_col = 1
        write_row += 1

        for col_num in range(1, source_network_day_sheet.max_column + 1):
            if source_network_day_sheet.cell(row=1, column=col_num).value in ['num_spots', 'equiv_units',
                                                                              'total_impressions',
                                                                              'total_unequiv_impressions',
                                                                              'total_unequiv_frequency', 'GRPs',
                                                                              'GRPs_unequiv',
                                                                              'target_impressions',
                                                                              'target_unequiv_impressions',
                                                                              'target_unequiv_frequency', 'TRPs',
                                                                              'TRPs_unequiv',
                                                                              'target_index_impressions',
                                                                              'target_index_unequiv_impressions',
                                                                              'tCPM']:
                if equiv:
                    if source_network_day_sheet.cell(row=1, column=col_num).value in ['equiv_units',
                                                                                      'total_impressions', 'GRPs',
                                                                                      'target_impressions', 'TRPs',
                                                                                      'target_index_impressions',
                                                                                      'tCPM']:
                        dest_network_day_sheet.cell(row=write_row,
                                                    column=write_col).value = source_network_day_sheet.cell(row=row_num,
                                                                                                            column=col_num).value
                        format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                                    source_network_day_sheet.cell(row=1, column=col_num).value)
                        write_col += 1

                    if source_network_day_sheet.cell(row=1, column=col_num).value == 'total_unequiv_frequency':
                        if row_num == 1:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = 'total_frequency'
                        else:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = \
                                source_network_day_sheet.cell(row=row_num,
                                                              column=6).value / source_network_day_sheet.cell(
                                    row=row_num,
                                    column=8).value

                        format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                                    source_network_day_sheet.cell(row=1, column=col_num).value)
                        write_col += 1

                    if source_network_day_sheet.cell(row=1, column=col_num).value == 'target_unequiv_frequency':
                        if row_num == 1:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = 'target_frequency'
                        else:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = \
                                source_network_day_sheet.cell(row=row_num,
                                                              column=18).value / source_network_day_sheet.cell(
                                    row=row_num,
                                    column=20).value
                        format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                                    source_network_day_sheet.cell(row=1, column=col_num).value)
                        write_col += 1

                else:
                    if source_network_day_sheet.cell(row=1, column=col_num).value in ['num_spots',
                                                                                      'total_unequiv_impressions',
                                                                                      'GRPs_unequiv',
                                                                                      'target_unequiv_impressions',
                                                                                      'TRPs_unequiv',
                                                                                      'target_index_unequiv_impressions',
                                                                                      'total_unequiv_frequency',
                                                                                      'target_unequiv_frequency']:
                        dest_network_day_sheet.cell(row=write_row,
                                                    column=write_col).value = source_network_day_sheet.cell(row=row_num,
                                                                                                            column=col_num).value
                        format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                                    source_network_day_sheet.cell(row=1, column=col_num).value)
                        write_col += 1

                    if source_network_day_sheet.cell(row=1, column=col_num).value == 'tCPM':
                        if row_num == 1:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = 'tCPM_unequiv'
                        else:
                            dest_network_day_sheet.cell(row=write_row, column=write_col).value = \
                                source_network_day_sheet.cell(row=row_num,
                                                              column=5).value * 1000 / source_network_day_sheet.cell(
                                    row=row_num,
                                    column=19).value
                            format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                                        source_network_day_sheet.cell(row=1, column=col_num).value)
                        write_col += 1

            else:
                dest_network_day_sheet.cell(row=write_row, column=write_col).value = source_network_day_sheet.cell(
                    row=row_num,
                    column=col_num).value

                format_cell(dest_network_day_sheet.cell(row=write_row, column=write_col),
                            source_network_day_sheet.cell(row=1, column=col_num).value)
                write_col += 1

    summary_wb.save(filename)

    return True


def process_frequency_distribution_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_freq_sheet = source_wb.get_sheet_by_name('Frequency Distribution')
    dest_freq_sheet = summary_wb.get_sheet_by_name('Frequency Distribution')
    dest_row = 4
    source_rows = source_freq_sheet.max_row

    dest_freq_sheet['A1'].font = Font(bold=True)
    dest_freq_sheet['A1'] = 'EQUIVALIZED - Frequency Distribution' if equiv else 'UNEQUIVALIZED - Frequency Distribution'

    for row_num in range(1, source_rows + 1):
        if source_freq_sheet.cell(row=row_num, column=1).value == 'Spot' and source_freq_sheet.cell(row=row_num,
                                                                                                    column=2).value == 'Total':
            for col_num in range(1, source_freq_sheet.max_column + 1):
                if source_freq_sheet.cell(row=1, column=col_num).value == 'frequency':
                    dest_freq_sheet.cell(row=dest_row, column=1).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=1), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'target':
                    dest_freq_sheet.cell(row=dest_row, column=3).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=3), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'total':
                    dest_freq_sheet.cell(row=dest_row, column=2).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=2), '')
            dest_row += 1

    dest_row += 1
    dest_freq_sheet.cell(row=dest_row, column=1).value = '# Networks'
    dest_freq_sheet.cell(row=dest_row, column=2).value = 'Total (Campaign)'
    dest_freq_sheet.cell(row=dest_row, column=3).value = 'Target (Campaign)'
    dest_freq_sheet.cell(row=dest_row, column=4).value = 'Target Composition'
    dest_row += 1

    for row_num in range(1, source_rows + 1):
        if source_freq_sheet.cell(row=row_num, column=1).value == 'Network' and source_freq_sheet.cell(row=row_num,
                                                                                                       column=2).value == 'Total':
            for col_num in range(1, source_freq_sheet.max_column + 1):
                if source_freq_sheet.cell(row=1, column=col_num).value == 'frequency':
                    dest_freq_sheet.cell(row=dest_row, column=1).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=1), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'target':
                    dest_freq_sheet.cell(row=dest_row, column=3).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=3), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'total':
                    dest_freq_sheet.cell(row=dest_row, column=2).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=2), '')
            dest_row += 1

    dest_freq_sheet.cell(row=dest_row, column=1).value = 'Total'
    dest_row += 2
    dest_freq_sheet.cell(row=dest_row, column=1).value = '# Programs'
    dest_freq_sheet.cell(row=dest_row, column=2).value = 'Total (Campaign)'
    dest_freq_sheet.cell(row=dest_row, column=3).value = 'Target (Campaign)'
    dest_row += 1

    for row_num in range(1, source_rows + 1):
        if source_freq_sheet.cell(row=row_num, column=1).value == 'Program' and source_freq_sheet.cell(row=row_num,
                                                                                                       column=2).value == 'Total':
            for col_num in range(1, source_freq_sheet.max_column + 1):
                if source_freq_sheet.cell(row=1, column=col_num).value == 'frequency':
                    dest_freq_sheet.cell(row=dest_row, column=1).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=1), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'target':
                    dest_freq_sheet.cell(row=dest_row, column=3).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=3), '')
                if source_freq_sheet.cell(row=1, column=col_num).value == 'total':
                    dest_freq_sheet.cell(row=dest_row, column=2).value = source_freq_sheet.cell(row=row_num,
                                                                                                column=col_num).value
                    format_cell(dest_freq_sheet.cell(row=dest_row, column=2), '')
            dest_row += 1

    # Calculations
    incrementor = 1
    sum_incrementor = 0
    network_start = 0
    for row_num in range(4, dest_freq_sheet.max_row + 1):
        if incrementor % 5 == 0 or not dest_freq_sheet.cell(row=row_num + 1, column=1).value:
            if dest_freq_sheet.cell(row=row_num, column=3).value:
                dest_freq_sheet.cell(row=row_num, column=4).value = sum_incrementor + dest_freq_sheet.cell(row=row_num,
                                                                                                           column=3).value
            else:
                dest_freq_sheet.cell(row=row_num, column=4).value = sum_incrementor

            format_cell(dest_freq_sheet.cell(row=row_num, column=4), '')
            incrementor = 1
            sum_incrementor = 0
        else:
            if dest_freq_sheet.cell(row=row_num, column=3).value:
                sum_incrementor += dest_freq_sheet.cell(row=row_num, column=3).value
            incrementor += 1
        if not dest_freq_sheet.cell(row=row_num + 1, column=1).value:
            network_start = row_num + 3
            break

    network_total = 0
    network_target = 0
    for row_num in range(network_start, dest_freq_sheet.max_row + 1):
        if dest_freq_sheet.cell(row=row_num, column=1).value != 'Total':
            network_total += dest_freq_sheet.cell(row=row_num, column=2).value
            network_target += dest_freq_sheet.cell(row=row_num, column=3).value if dest_freq_sheet.cell(row=row_num,
                                                                                                        column=3).value else 0
        else:
            dest_freq_sheet.cell(row=row_num, column=2).value = network_total
            format_cell(dest_freq_sheet.cell(row=row_num, column=2), '')
            dest_freq_sheet.cell(row=row_num, column=3).value = network_target
            format_cell(dest_freq_sheet.cell(row=row_num, column=3), '')
            break

    for row_num in range(network_start, dest_freq_sheet.max_row + 1):
        if dest_freq_sheet.cell(row=row_num, column=1).value != 'Total':
            dest_freq_sheet.cell(row=row_num, column=4).value = dest_freq_sheet.cell(row=row_num,
                                                                                     column=3).value / network_target if dest_freq_sheet.cell(
                row=row_num, column=3).value else 0
            format_cell(dest_freq_sheet.cell(row=row_num, column=4), 'pct')
        else:
            break

    summary_wb.save(filename)

    return True


def process_reach_by_week_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_reach_by_week_sheet = source_wb.get_sheet_by_name('Reach by Week')
    dest_reach_by_week_sheet = summary_wb.get_sheet_by_name('Reach by Week')
    dest_row = 4
    write_column = 1

    equiv_list = [6, 11]
    unequiv_list = [7, 12]
    equiv_calc_list = [8,13]

    dest_reach_by_week_sheet['A1'].font = Font(bold=True)
    dest_reach_by_week_sheet['A1'] = 'EQUIVALIZED - Reach By Week' if equiv else 'UNEQUIVALIZED - Reach By Week'

    for row_num in range(2, source_reach_by_week_sheet.max_row + 1):
        if source_reach_by_week_sheet.cell(row=row_num, column=1).value == 'Total':
            for col_num in range(2, source_reach_by_week_sheet.max_column + 1):
                if not equiv and col_num in equiv_list:
                    continue
                if equiv and col_num in unequiv_list:
                    continue

                if equiv and col_num in equiv_calc_list:
                    if col_num == 8:
                        dest_reach_by_week_sheet.cell(row=dest_row, column=write_column).value = source_reach_by_week_sheet.cell(row=row_num, column=6).value / source_reach_by_week_sheet.cell(row=row_num, column=4).value
                        format_cell(dest_reach_by_week_sheet.cell(row=dest_row, column=write_column),source_reach_by_week_sheet.cell(row=1, column=col_num).value)
                    if col_num == 13:
                        dest_reach_by_week_sheet.cell(row=dest_row, column=write_column).value = source_reach_by_week_sheet.cell(row=row_num,column=11).value / source_reach_by_week_sheet.cell(row=row_num, column=9).value
                        format_cell(dest_reach_by_week_sheet.cell(row=dest_row, column=write_column),source_reach_by_week_sheet.cell(row=1, column=col_num).value)
                    write_column += 1
                    continue

                dest_reach_by_week_sheet.cell(row=dest_row,
                                              column=write_column).value = source_reach_by_week_sheet.cell(row=row_num,
                                                                                                           column=col_num).value
                format_cell(dest_reach_by_week_sheet.cell(row=dest_row, column=write_column),source_reach_by_week_sheet.cell(row=1, column=col_num).value)
                write_column += 1

            dest_row += 1
            write_column = 1

    summary_wb.save(filename)
    return True


def process_frequency_distribution_by_net_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_freq_sheet = source_wb.get_sheet_by_name('Frequency Distribution')
    dest_freq_sheet = summary_wb.get_sheet_by_name('Freq Distribution by Net')
    source_rows = source_freq_sheet.max_row
    freq_obj = {}

    dest_freq_sheet['A1'].font = Font(bold=True)
    dest_freq_sheet['A1'] = 'EQUIVALIZED - Freq Distribution by Net' if equiv else 'UNEQUIVALIZED - Freq Distribution by Net'

    for row_num in range(2, source_rows + 1):
        if freq_obj.has_key(int(source_freq_sheet.cell(row=row_num, column=3).value)):
            if freq_obj[int(source_freq_sheet.cell(row=row_num, column=3).value)].has_key(
                    source_freq_sheet.cell(row=row_num, column=1).value):
                freq_obj[int(source_freq_sheet.cell(row=row_num, column=3).value)][
                    source_freq_sheet.cell(row=row_num, column=1).value].append(
                    {source_freq_sheet.cell(row=row_num, column=2).value:
                         {'target': source_freq_sheet.cell(row=row_num, column=4).value,
                          'total': source_freq_sheet.cell(row=row_num, column=5).value}}
                )
            else:
                freq_obj[int(source_freq_sheet.cell(row=row_num, column=3).value)][
                    source_freq_sheet.cell(row=row_num, column=1).value] = [
                    {source_freq_sheet.cell(row=row_num, column=2).value:
                         {'target': source_freq_sheet.cell(row=row_num, column=4).value,
                          'total': source_freq_sheet.cell(row=row_num, column=5).value}}
                ]
        else:
            freq_obj[int(source_freq_sheet.cell(row=row_num, column=3).value)] = {
                source_freq_sheet.cell(row=row_num, column=1).value: [{
                    source_freq_sheet.cell(row=row_num, column=2).value:
                        {'target': source_freq_sheet.cell(row=row_num, column=4).value,
                         'total': source_freq_sheet.cell(row=row_num, column=5).value}}]}

    write_row = 4
    for freq, row_list in collections.OrderedDict(sorted(freq_obj.items())).items():
        dest_freq_sheet.cell(row=write_row, column=1).value = freq
        for row_obj in row_list['Spot']:
            if row_obj.has_key('Total'):
                dest_freq_sheet.cell(row=write_row, column=2).value = row_obj['Total']['total']
                dest_freq_sheet.cell(row=write_row, column=3).value = row_obj['Total']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=2), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=3), '')
            if row_obj.has_key('Bravo'):
                dest_freq_sheet.cell(row=write_row, column=4).value = row_obj['Bravo']['total']
                dest_freq_sheet.cell(row=write_row, column=5).value = row_obj['Bravo']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=4), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=5), '')
            if row_obj.has_key('CNBC'):
                dest_freq_sheet.cell(row=write_row, column=6).value = row_obj['CNBC']['total']
                dest_freq_sheet.cell(row=write_row, column=7).value = row_obj['CNBC']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=6), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=7), '')
            if row_obj.has_key('E!'):
                dest_freq_sheet.cell(row=write_row, column=8).value = row_obj['E!']['total']
                dest_freq_sheet.cell(row=write_row, column=9).value = row_obj['E!']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=8), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=9), '')
            if row_obj.has_key('Esquire'):
                dest_freq_sheet.cell(row=write_row, column=10).value = row_obj['Esquire']['total']
                dest_freq_sheet.cell(row=write_row, column=11).value = row_obj['Esquire']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=10), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=11), '')
            if row_obj.has_key('Golf Channel'):
                dest_freq_sheet.cell(row=write_row, column=12).value = row_obj['Golf Channel']['total']
                dest_freq_sheet.cell(row=write_row, column=13).value = row_obj['Golf Channel']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=12), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=13), '')
            if row_obj.has_key('MSNBC'):
                dest_freq_sheet.cell(row=write_row, column=14).value = row_obj['MSNBC']['total']
                dest_freq_sheet.cell(row=write_row, column=15).value = row_obj['MSNBC']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=14), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=15), '')
            if row_obj.has_key('NBC'):
                dest_freq_sheet.cell(row=write_row, column=16).value = row_obj['NBC']['total']
                dest_freq_sheet.cell(row=write_row, column=17).value = row_obj['NBC']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=16), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=17), '')
            if row_obj.has_key('NBCSN'):
                dest_freq_sheet.cell(row=write_row, column=18).value = row_obj['NBCSN']['total']
                dest_freq_sheet.cell(row=write_row, column=19).value = row_obj['NBCSN']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=18), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=19), '')
            if row_obj.has_key('Oxygen'):
                dest_freq_sheet.cell(row=write_row, column=20).value = row_obj['Oxygen']['total']
                dest_freq_sheet.cell(row=write_row, column=21).value = row_obj['Oxygen']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=20), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=21), '')
            if row_obj.has_key('Sprout'):
                dest_freq_sheet.cell(row=write_row, column=22).value = row_obj['Sprout']['total']
                dest_freq_sheet.cell(row=write_row, column=23).value = row_obj['Sprout']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=22), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=23), '')
            if row_obj.has_key('Syfy'):
                dest_freq_sheet.cell(row=write_row, column=24).value = row_obj['Syfy']['total']
                dest_freq_sheet.cell(row=write_row, column=25).value = row_obj['Syfy']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=24), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=25), '')
            if row_obj.has_key('Telemundo'):
                dest_freq_sheet.cell(row=write_row, column=26).value = row_obj['Telemundo']['total']
                dest_freq_sheet.cell(row=write_row, column=27).value = row_obj['Telemundo']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=26), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=27), '')
            if row_obj.has_key('USA'):
                dest_freq_sheet.cell(row=write_row, column=28).value = row_obj['USA']['total']
                dest_freq_sheet.cell(row=write_row, column=29).value = row_obj['USA']['target']
                format_cell(dest_freq_sheet.cell(row=write_row, column=28), '')
                format_cell(dest_freq_sheet.cell(row=write_row, column=29), '')
        write_row += 1

    write_row += 1
    dest_freq_sheet.cell(row=write_row, column=1).value = '# Programs'
    for col_num in range(2, dest_freq_sheet.max_column + 1):
        dest_freq_sheet.cell(row=write_row, column=col_num).value = dest_freq_sheet.cell(row=1, column=col_num).value
    write_row += 1

    for freq, row_list in collections.OrderedDict(freq_obj.items()).items():
        dest_freq_sheet.cell(row=write_row, column=1).value = freq
        if row_list.has_key('Program'):
            for row_obj in row_list['Program']:
                if row_obj.has_key('Total'):
                    dest_freq_sheet.cell(row=write_row, column=2).value = row_obj['Total']['total']
                    dest_freq_sheet.cell(row=write_row, column=3).value = row_obj['Total']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=2), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=3), '')
                if row_obj.has_key('Bravo'):
                    dest_freq_sheet.cell(row=write_row, column=4).value = row_obj['Bravo']['total']
                    dest_freq_sheet.cell(row=write_row, column=5).value = row_obj['Bravo']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=4), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=5), '')
                if row_obj.has_key('CNBC'):
                    dest_freq_sheet.cell(row=write_row, column=6).value = row_obj['CNBC']['total']
                    dest_freq_sheet.cell(row=write_row, column=7).value = row_obj['CNBC']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=6), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=7), '')
                if row_obj.has_key('E!'):
                    dest_freq_sheet.cell(row=write_row, column=8).value = row_obj['E!']['total']
                    dest_freq_sheet.cell(row=write_row, column=9).value = row_obj['E!']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=8), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=9), '')
                if row_obj.has_key('Esquire'):
                    dest_freq_sheet.cell(row=write_row, column=10).value = row_obj['Esquire']['total']
                    dest_freq_sheet.cell(row=write_row, column=11).value = row_obj['Esquire']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=10), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=11), '')
                if row_obj.has_key('Golf Channel'):
                    dest_freq_sheet.cell(row=write_row, column=12).value = row_obj['Golf Channel']['total']
                    dest_freq_sheet.cell(row=write_row, column=13).value = row_obj['Golf Channel']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=12), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=13), '')
                if row_obj.has_key('MSNBC'):
                    dest_freq_sheet.cell(row=write_row, column=14).value = row_obj['MSNBC']['total']
                    dest_freq_sheet.cell(row=write_row, column=15).value = row_obj['MSNBC']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=14), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=15), '')
                if row_obj.has_key('NBC'):
                    dest_freq_sheet.cell(row=write_row, column=16).value = row_obj['NBC']['total']
                    dest_freq_sheet.cell(row=write_row, column=17).value = row_obj['NBC']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=16), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=17), '')
                if row_obj.has_key('NBCSN'):
                    dest_freq_sheet.cell(row=write_row, column=18).value = row_obj['NBCSN']['total']
                    dest_freq_sheet.cell(row=write_row, column=19).value = row_obj['NBCSN']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=18), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=19), '')
                if row_obj.has_key('Oxygen'):
                    dest_freq_sheet.cell(row=write_row, column=20).value = row_obj['Oxygen']['total']
                    dest_freq_sheet.cell(row=write_row, column=21).value = row_obj['Oxygen']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=20), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=21), '')
                if row_obj.has_key('Sprout'):
                    dest_freq_sheet.cell(row=write_row, column=22).value = row_obj['Sprout']['total']
                    dest_freq_sheet.cell(row=write_row, column=23).value = row_obj['Sprout']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=22), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=23), '')
                if row_obj.has_key('Syfy'):
                    dest_freq_sheet.cell(row=write_row, column=24).value = row_obj['Syfy']['total']
                    dest_freq_sheet.cell(row=write_row, column=25).value = row_obj['Syfy']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=24), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=25), '')
                if row_obj.has_key('Telemundo'):
                    dest_freq_sheet.cell(row=write_row, column=26).value = row_obj['Telemundo']['total']
                    dest_freq_sheet.cell(row=write_row, column=27).value = row_obj['Telemundo']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=26), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=27), '')
                if row_obj.has_key('USA'):
                    dest_freq_sheet.cell(row=write_row, column=28).value = row_obj['USA']['total']
                    dest_freq_sheet.cell(row=write_row, column=29).value = row_obj['USA']['target']
                    format_cell(dest_freq_sheet.cell(row=write_row, column=28), '')
                    format_cell(dest_freq_sheet.cell(row=write_row, column=29), '')
            write_row += 1

    summary_wb.save(filename)

    return True


def process_network_reach_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_reach_net_sheet = source_wb.get_sheet_by_name('Reach by Week')
    dest_reach_net_sheet = summary_wb.get_sheet_by_name('Network Reach by Week')
    source_rows = source_reach_net_sheet.max_row
    reach_net_obj = {}

    dest_reach_net_sheet['A1'].font = Font(bold=True)
    dest_reach_net_sheet['A1'] = 'EQUIVALIZED - Network Reach By Week' if equiv else 'UNEQUIVALIZED - Network Reach By Week'

    for row_num in range(2, source_rows + 1):
        if reach_net_obj.has_key(int(source_reach_net_sheet.cell(row=row_num, column=2).value)):
            reach_net_obj[int(source_reach_net_sheet.cell(row=row_num, column=2).value)][
                source_reach_net_sheet.cell(row=row_num, column=1).value] = \
                {'weekof': source_reach_net_sheet.cell(row=row_num, column=3).value,
                 'total': source_reach_net_sheet.cell(row=row_num, column=4).value,
                 'total_pct': source_reach_net_sheet.cell(row=row_num, column=5).value,
                 'total_impressions': source_reach_net_sheet.cell(row=row_num, column=6).value,
                 'total_impressions_unequiv': source_reach_net_sheet.cell(row=row_num, column=7).value,
                 'total_frequency_unequiv': source_reach_net_sheet.cell(row=row_num, column=8).value,
                 'target': source_reach_net_sheet.cell(row=row_num, column=9).value,
                 'target_pct': source_reach_net_sheet.cell(row=row_num, column=10).value,
                 'target_impressions': source_reach_net_sheet.cell(row=row_num, column=11).value,
                 'target_impressions_unequiv': source_reach_net_sheet.cell(row=row_num, column=12).value,
                 'target_frequency_unequiv': source_reach_net_sheet.cell(row=row_num, column=13).value}
        else:
            reach_net_obj[int(source_reach_net_sheet.cell(row=row_num, column=2).value)] = {
                source_reach_net_sheet.cell(row=row_num, column=1).value:
                    {'weekof': source_reach_net_sheet.cell(row=row_num, column=3).value,
                     'total': source_reach_net_sheet.cell(row=row_num, column=4).value,
                     'total_pct': source_reach_net_sheet.cell(row=row_num, column=5).value,
                     'total_impressions': source_reach_net_sheet.cell(row=row_num, column=6).value,
                     'total_impressions_unequiv': source_reach_net_sheet.cell(row=row_num, column=7).value,
                     'total_frequency_unequiv': source_reach_net_sheet.cell(row=row_num, column=8).value,
                     'target': source_reach_net_sheet.cell(row=row_num, column=9).value,
                     'target_pct': source_reach_net_sheet.cell(row=row_num, column=10).value,
                     'target_impressions': source_reach_net_sheet.cell(row=row_num, column=11).value,
                     'target_impressions_unequiv': source_reach_net_sheet.cell(row=row_num, column=12).value,
                     'target_frequency_unequiv': source_reach_net_sheet.cell(row=row_num, column=13).value}}
    write_row = 4
    for week, row_list in collections.OrderedDict(sorted(reach_net_obj.items())).items():
        dest_reach_net_sheet.cell(row=write_row, column=1).value = row_list['Total']['weekof']
        dest_reach_net_sheet.cell(row=write_row, column=2).value = row_list['Total']['target_pct']
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=2), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=3).value = row_list['NBC']['target_pct'] if row_list.has_key(
            'NBC') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=3), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=5).value = row_list['Total']['weekof']
        dest_reach_net_sheet.cell(row=write_row, column=6).value = row_list['Bravo']['target_pct'] if row_list.has_key(
            'Bravo') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=6), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=7).value = row_list['Chiller']['target_pct'] if row_list.has_key(
            'Chiller') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=7), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=8).value = row_list['CNBC'][
            'target_pct'] if row_list.has_key('CNBC') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=8), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=9).value = row_list['E!']['target_pct'] if row_list.has_key(
            'E!') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=9), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=10).value = row_list['Esquire'][
            'target_pct'] if row_list.has_key('Esquire') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=10), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=11).value = row_list['Golf Channel'][
            'target_pct'] if row_list.has_key('Golf Channel') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=11), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=12).value = row_list['MSNBC']['target_pct'] if row_list.has_key(
            'MSNBC') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=12), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=13).value = row_list['NBCSN']['target_pct'] if row_list.has_key(
            'NBCSN') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=13), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=14).value = row_list['Oxygen'][
            'target_pct'] if row_list.has_key('Oxygen') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=14), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=15).value = row_list['Sprout']['target_pct'] if row_list.has_key(
            'Sprout') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=15), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=16).value = row_list['Syfy']['target_pct'] if row_list.has_key(
            'Syfy') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=16), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=17).value = row_list['Telemundo']['target_pct'] if row_list.has_key(
            'Telemundo') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=17), 'pct')
        dest_reach_net_sheet.cell(row=write_row, column=18).value = row_list['USA']['target_pct'] if row_list.has_key(
            'USA') else ''
        format_cell(dest_reach_net_sheet.cell(row=write_row, column=18), 'pct')
        write_row += 1

    networks = ['Bravo','Chiller','CNBC','E!','Esquire','Golf Channel','MSNBC','NBC','NBCSN','Oxygen','Sprout','Syfy','Telemundo','USA']
    
    for network in networks:
        if row_list.has_key(network):
            write_row += 3
            dest_reach_net_sheet.cell(row=write_row, column=1).value = network
            write_row += 1
            dest_reach_net_sheet.cell(row=write_row, column=1).value = 'week'
            dest_reach_net_sheet.cell(row=write_row, column=2).value = 'week_of'
            dest_reach_net_sheet.cell(row=write_row, column=3).value = 'reach_total'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=3), '')
            dest_reach_net_sheet.cell(row=write_row, column=4).value = 'reach_pct_total'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=4), 'pct')
            if equiv:
                dest_reach_net_sheet.cell(row=write_row, column=5).value = 'impressions_total'
                dest_reach_net_sheet.cell(row=write_row, column=6).value = 'avg_freq_total'
            else:
                dest_reach_net_sheet.cell(row=write_row, column=5).value = 'impressions_unequiv_total'
                dest_reach_net_sheet.cell(row=write_row, column=6).value = 'avg_freq_unequiv_total'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=5), '')
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=6), '')
            dest_reach_net_sheet.cell(row=write_row, column=7).value = 'reach_target'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=7), '')
            dest_reach_net_sheet.cell(row=write_row, column=8).value = 'reach_pct_target'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=8), 'pct')
            if equiv:
                dest_reach_net_sheet.cell(row=write_row, column=9).value = 'impressions_target'
                dest_reach_net_sheet.cell(row=write_row, column=10).value = 'avg_freq_target'
            else:
                dest_reach_net_sheet.cell(row=write_row, column=9).value = 'impressions_unequiv_target'
                dest_reach_net_sheet.cell(row=write_row, column=10).value = 'avg_freq_unequiv_target'
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=9), '')
            format_cell(dest_reach_net_sheet.cell(row=write_row, column=10), '')

            write_row += 1

            for week, row_list in collections.OrderedDict(sorted(reach_net_obj.items())).items():
                dest_reach_net_sheet.cell(row=write_row, column=1).value = week
                if row_list.has_key(network):
                    dest_reach_net_sheet.cell(row=write_row, column=2).value = row_list[network]['weekof']
                    dest_reach_net_sheet.cell(row=write_row, column=3).value = row_list[network]['total']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=3), '')
                    dest_reach_net_sheet.cell(row=write_row, column=4).value = row_list[network]['total_pct']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=4), 'pct')
                    if equiv:
                        dest_reach_net_sheet.cell(row=write_row, column=5).value = row_list[network]['total_impressions']
                        dest_reach_net_sheet.cell(row=write_row, column=6).value = row_list[network]['total_impressions'] / \
                                                                                   row_list[network]['total']
                    else:
                        dest_reach_net_sheet.cell(row=write_row, column=5).value = row_list[network][
                            'total_impressions_unequiv']
                        dest_reach_net_sheet.cell(row=write_row, column=6).value = row_list[network][
                            'total_frequency_unequiv']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=5), '')
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=6), '')

                    dest_reach_net_sheet.cell(row=write_row, column=7).value = row_list[network]['target']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=7), '')
                    dest_reach_net_sheet.cell(row=write_row, column=8).value = row_list[network]['target_pct']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=8), 'pct')
                    if equiv:
                        dest_reach_net_sheet.cell(row=write_row, column=9).value = row_list[network]['target_impressions']
                        dest_reach_net_sheet.cell(row=write_row, column=10).value = row_list[network][
                                                                                        'target_impressions'] / \
                                                                                    row_list[network]['target']
                    else:
                        dest_reach_net_sheet.cell(row=write_row, column=9).value = row_list[network][
                            'target_impressions_unequiv']
                        dest_reach_net_sheet.cell(row=write_row, column=10).value = row_list[network][
                            'target_frequency_unequiv']
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=9), '')
                    format_cell(dest_reach_net_sheet.cell(row=write_row, column=10), '')

                    write_row += 1

    summary_wb.save(filename)

    return True


def process_powerpoint_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_sheet = summary_wb.get_sheet_by_name('Summary Metrics')
    dest_pp_sheet = summary_wb.get_sheet_by_name('Powerpoint Data')
    source_rows = source_sheet.max_row

    dest_pp_sheet['A2'] = 'EQUIVALIZED - Total Campaign Target Metrics' if equiv else 'UNEQUIVALIZED - Total Campaign Target Metrics'

    if 'normalized' in filename:
        dest_column = 2
    else:
        dest_column = 3

    for row_num in range(1, source_rows):
        if source_sheet.cell(row=row_num, column=1).value == 'Total':
            dest_pp_sheet.cell(row=4, column=dest_column).value = source_sheet.cell(row=row_num, column=3).value
            format_cell(dest_pp_sheet.cell(row=4, column=dest_column), 'dollars')
            dest_pp_sheet.cell(row=5, column=dest_column).value = source_sheet.cell(row=row_num, column=2).value
            dest_pp_sheet.cell(row=8, column=dest_column).value = source_sheet.cell(row=row_num, column=14).value
            format_cell(dest_pp_sheet.cell(row=8, column=dest_column), '')
            dest_pp_sheet.cell(row=9, column=dest_column).value = source_sheet.cell(row=row_num, column=26).value
            format_cell(dest_pp_sheet.cell(row=9, column=dest_column), '')
            dest_pp_sheet.cell(row=10, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
            format_cell(dest_pp_sheet.cell(row=10, column=dest_column), 'pct')
            dest_pp_sheet.cell(row=11, column=dest_column).value = source_sheet.cell(row=row_num, column=23).value
            format_cell(dest_pp_sheet.cell(row=11, column=dest_column), '')
            dest_pp_sheet.cell(row=12, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            format_cell(dest_pp_sheet.cell(row=12, column=dest_column), '')
            dest_pp_sheet.cell(row=13, column=dest_column).value = source_sheet.cell(row=row_num, column=22).value
            format_cell(dest_pp_sheet.cell(row=13, column=dest_column), '')
            dest_pp_sheet.cell(row=14, column=dest_column).value = source_sheet.cell(row=row_num, column=15).value
            format_cell(dest_pp_sheet.cell(row=14, column=dest_column), '')
            dest_pp_sheet.cell(row=45, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            format_cell(dest_pp_sheet.cell(row=45, column=dest_column), '')
            dest_pp_sheet.cell(row=72, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
            format_cell(dest_pp_sheet.cell(row=72, column=dest_column), 'pct')
            break

    start_row = 0
    end_row = 0
    for row_num in range(2, source_rows):
        if source_sheet.cell(row=row_num, column=1).value == 'Network Summary':
            start_row = row_num

        if not source_sheet.cell(row=row_num, column=1).value and start_row:
            end_row = row_num + 1

    for row_num in range(start_row, end_row):
        if source_sheet.cell(row=row_num, column=1).value == 'NBC':
            dest_pp_sheet.cell(row=25, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=53, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=80, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Bravo':
            dest_pp_sheet.cell(row=18, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=46, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=73, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Chiller':
            dest_pp_sheet.cell(row=19, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=47, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=74, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'CNBC':
            dest_pp_sheet.cell(row=20, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=48, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=75, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'E!':
            dest_pp_sheet.cell(row=21, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=49, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=76, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Esquire':
            dest_pp_sheet.cell(row=22, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=50, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=77, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Golf Channel':
            dest_pp_sheet.cell(row=23, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=51, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=78, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'MSNBC':
            dest_pp_sheet.cell(row=24, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=52, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=79, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBCSN':
            dest_pp_sheet.cell(row=26, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=54, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=81, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Oxygen':
            dest_pp_sheet.cell(row=27, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=55, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=82, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Sprout':
            dest_pp_sheet.cell(row=28, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=56, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=83, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Syfy':
            dest_pp_sheet.cell(row=29, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=57, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=84, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'Telemundo':
            dest_pp_sheet.cell(row=30, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=58, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=85, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value
        if source_sheet.cell(row=row_num, column=1).value == 'USA':
            dest_pp_sheet.cell(row=31, column=dest_column).value = source_sheet.cell(row=row_num,
                                                                           column=14).value / dest_pp_sheet.cell(
                row=8, column=dest_column).value
            dest_pp_sheet.cell(row=59, column=dest_column).value = source_sheet.cell(row=row_num, column=24).value
            dest_pp_sheet.cell(row=86, column=dest_column).value = source_sheet.cell(row=row_num, column=16).value

    source_sheet = summary_wb.get_sheet_by_name('Program Metrics')
    source_rows = source_sheet.max_row

    programs_obj = {}
    networks_obj = {}
    for row_num in range(4, source_rows + 1):
        networks_obj[source_sheet.cell(row=row_num, column=1).value] = 1
        programs_obj[source_sheet.cell(row=row_num, column=18).value] = 1

    dest_pp_sheet.cell(row=6, column=dest_column).value = len(programs_obj.items())
    dest_pp_sheet.cell(row=7, column=dest_column).value = len(networks_obj.items())

    source_sheet = summary_wb.get_sheet_by_name('Network Daypart')
    source_rows = source_sheet.max_row

    nbc_sum = 0
    nbc_morning = 0
    nbc_daytime = 0
    nbc_early = 0
    nbc_prime = 0
    nbc_late = 0
    nbc_overnight = 0
    for row_num in range(4, source_rows + 1):
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Morning':
            nbc_morning = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=63, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=90, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Daytime':
            nbc_daytime = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=64, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=91, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Early Fringe':
            nbc_early = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=65, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=92, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Prime':
            nbc_prime = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=66, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=93, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Late Night':
            nbc_late = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=67, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=94, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value
        if source_sheet.cell(row=row_num, column=1).value == 'NBC' and source_sheet.cell(row=row_num,
                                                                                         column=2).value == 'Overnight':
            nbc_overnight = source_sheet.cell(row=row_num, column=15).value
            nbc_sum += source_sheet.cell(row=row_num, column=15).value
            dest_pp_sheet.cell(row=68, column=dest_column).value = source_sheet.cell(row=row_num, column=25).value
            dest_pp_sheet.cell(row=95, column=dest_column).value = source_sheet.cell(row=row_num, column=17).value

    dest_pp_sheet.cell(row=36, column=dest_column).value = nbc_morning / nbc_sum if nbc_sum > 0 else 0
    dest_pp_sheet.cell(row=37, column=dest_column).value = nbc_daytime / nbc_sum if nbc_sum > 0 else 0
    dest_pp_sheet.cell(row=38, column=dest_column).value = nbc_early / nbc_sum if nbc_sum > 0 else 0
    dest_pp_sheet.cell(row=39, column=dest_column).value = nbc_prime / nbc_sum if nbc_sum > 0 else 0
    dest_pp_sheet.cell(row=40, column=dest_column).value = nbc_late / nbc_sum if nbc_sum > 0 else 0
    dest_pp_sheet.cell(row=41, column=dest_column).value = nbc_overnight / nbc_sum if nbc_sum > 0 else 0

    summary_wb._active_sheet_index = 6

    summary_wb.save(filename)
    return True


def process_appendix_tab(filename, equiv):
    global source_wb
    global summary_wb

    source_sheet = summary_wb.get_sheet_by_name('Network Daypart')
    dest_pp_sheet = summary_wb.get_sheet_by_name('Appendix')
    source_rows = source_sheet.max_row
    day_net_obj = {'Morning': {}, 'Daytime': {}, 'Early Fringe': {}, 'Prime': {}, 'Late Night': {}, 'Overnight': {}}
    dest_pp_sheet['A1'] = 'EQUIVALIZED - Appendix' if equiv else 'UNEQUIVALIZED - Appendix'

    for row_num in range(4, source_rows + 1):
        if day_net_obj.has_key(source_sheet.cell(row=row_num, column=2).value):
            day_net_obj[source_sheet.cell(row=row_num, column=2).value][
                source_sheet.cell(row=row_num, column=1).value] = \
                {'target_impressions': source_sheet.cell(row=row_num, column=15).value,
                 'target_index': source_sheet.cell(row=row_num, column=25).value,
                 'target_reach': source_sheet.cell(row=row_num, column=16).value,
                 'target_frequency': source_sheet.cell(row=row_num, column=23).value,
                 'tCPM': source_sheet.cell(row=row_num, column=27).value}
        else:
            day_net_obj[source_sheet.cell(row=row_num, column=2).value] = {
                source_sheet.cell(row=row_num, column=1).value:
                    {'target_impressions': source_sheet.cell(row=row_num, column=15).value,
                     'target_index': source_sheet.cell(row=row_num, column=25).value,
                     'target_reach': source_sheet.cell(row=row_num, column=16).value,
                     'target_frequency': source_sheet.cell(row=row_num, column=23).value,
                     'tCPM': source_sheet.cell(row=row_num, column=27).value}}

    source_sheet = source_wb.get_sheet_by_name('Summary')
    source_rows = source_sheet.max_row
    net_obj = {}

    for row_num in range(2, source_rows + 1):
        if equiv:
            net_obj[source_sheet.cell(row=row_num, column=1).value] = {
                'target_impressions': source_sheet.cell(row=row_num, column=17).value,
                'target_index': source_sheet.cell(row=row_num, column=29).value,
                'target_reach': source_sheet.cell(row=row_num, column=19).value,
                'target_frequency': source_sheet.cell(row=row_num, column=26).value,
                'tCPM': source_sheet.cell(row=row_num, column=4).value * 1000 / source_sheet.cell(row=row_num, column=17).value}
        else:
            net_obj[source_sheet.cell(row=row_num, column=1).value] = {
                'target_impressions': source_sheet.cell(row=row_num, column=18).value,
                'target_index': source_sheet.cell(row=row_num, column=30).value,
                'target_reach': source_sheet.cell(row=row_num, column=19).value,
                'target_frequency': source_sheet.cell(row=row_num, column=26).value,
                'tCPM': source_sheet.cell(row=row_num, column=4).value * 1000 / source_sheet.cell(row=row_num, column=18).value}
    if 'normalized' in filename:
        # Morning
        mo_im_total = 0
        mo_re_total = 0
        if day_net_obj['Morning'].has_key('Bravo'):
            dest_pp_sheet['B5'] = day_net_obj['Morning']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B28'] = day_net_obj['Morning']['Bravo']['target_index']
            dest_pp_sheet['B52'] = day_net_obj['Morning']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B76'] = day_net_obj['Morning']['Bravo']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Chiller'):
            dest_pp_sheet['C5'] = day_net_obj['Morning']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C28'] = day_net_obj['Morning']['Chiller']['target_index']
            dest_pp_sheet['C52'] = day_net_obj['Morning']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C76'] = day_net_obj['Morning']['Chiller']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('CNBC'):
            dest_pp_sheet['D5'] = day_net_obj['Morning']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D28'] = day_net_obj['Morning']['CNBC']['target_index']
            dest_pp_sheet['D52'] = day_net_obj['Morning']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D76'] = day_net_obj['Morning']['CNBC']['target_frequency']
            mo_im_total += day_net_obj['Morning']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('E!'):
            dest_pp_sheet['E5'] = day_net_obj['Morning']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E28'] = day_net_obj['Morning']['E!']['target_index']
            dest_pp_sheet['E52'] = day_net_obj['Morning']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E76'] = day_net_obj['Morning']['E!']['target_frequency']
            mo_im_total += day_net_obj['Morning']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Esquire'):
            dest_pp_sheet['F5'] = day_net_obj['Morning']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F28'] = day_net_obj['Morning']['Esquire']['target_index']
            dest_pp_sheet['F52'] = day_net_obj['Morning']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F76'] = day_net_obj['Morning']['Esquire']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Golf Channel'):
            dest_pp_sheet['G5'] = day_net_obj['Morning']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G28'] = day_net_obj['Morning']['Golf Channel']['target_index']
            dest_pp_sheet['G52'] = day_net_obj['Morning']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G76'] = day_net_obj['Morning']['Golf Channel']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('MSNBC'):
            dest_pp_sheet['H5'] = day_net_obj['Morning']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H28'] = day_net_obj['Morning']['MSNBC']['target_index']
            dest_pp_sheet['H52'] = day_net_obj['Morning']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            mo_im_total += day_net_obj['Morning']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('NBC'):
            dest_pp_sheet['I5'] = day_net_obj['Morning']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I28'] = day_net_obj['Morning']['NBC']['target_index']
            dest_pp_sheet['I52'] = day_net_obj['Morning']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I76'] = day_net_obj['Morning']['NBC']['target_frequency']
            mo_im_total += day_net_obj['Morning']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('NBCSN'):
            dest_pp_sheet['J5'] = day_net_obj['Morning']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J28'] = day_net_obj['Morning']['NBCSN']['target_index']
            dest_pp_sheet['J52'] = day_net_obj['Morning']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J76'] = day_net_obj['Morning']['NBCSN']['target_frequency']
            mo_im_total += day_net_obj['Morning']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Oxygen'):
            dest_pp_sheet['K5'] = day_net_obj['Morning']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K28'] = day_net_obj['Morning']['Oxygen']['target_index']
            dest_pp_sheet['K52'] = day_net_obj['Morning']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K76'] = day_net_obj['Morning']['Oxygen']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Sprout'):
            dest_pp_sheet['L5'] = day_net_obj['Morning']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L28'] = day_net_obj['Morning']['Sprout']['target_index']
            dest_pp_sheet['L52'] = day_net_obj['Morning']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L76'] = day_net_obj['Morning']['Sprout']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Syfy'):
            dest_pp_sheet['M5'] = day_net_obj['Morning']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M28'] = day_net_obj['Morning']['Syfy']['target_index']
            dest_pp_sheet['M52'] = day_net_obj['Morning']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M76'] = day_net_obj['Morning']['Syfy']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Telemundo'):
            dest_pp_sheet['N5'] = day_net_obj['Morning']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N28'] = day_net_obj['Morning']['Telemundo']['target_index']
            dest_pp_sheet['N52'] = day_net_obj['Morning']['Telemundo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['N76'] = day_net_obj['Morning']['Telemundo']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('USA'):
            dest_pp_sheet['O5'] = day_net_obj['Morning']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O28'] = day_net_obj['Morning']['USA']['target_index']
            dest_pp_sheet['O52'] = day_net_obj['Morning']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O76'] = day_net_obj['Morning']['USA']['target_frequency']
            mo_im_total += day_net_obj['Morning']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Daytime
        dy_im_total = 0
        dy_re_total = 0
        if day_net_obj['Daytime'].has_key('Bravo'):
            dest_pp_sheet['B6'] = day_net_obj['Daytime']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B29'] = day_net_obj['Daytime']['Bravo']['target_index']
            dest_pp_sheet['B53'] = day_net_obj['Daytime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B77'] = day_net_obj['Daytime']['Bravo']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Chiller'):
            dest_pp_sheet['C6'] = day_net_obj['Daytime']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C29'] = day_net_obj['Daytime']['Chiller']['target_index']
            dest_pp_sheet['C53'] = day_net_obj['Daytime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C77'] = day_net_obj['Daytime']['Chiller']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('CNBC'):
            dest_pp_sheet['D6'] = day_net_obj['Daytime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D29'] = day_net_obj['Daytime']['CNBC']['target_index']
            dest_pp_sheet['D53'] = day_net_obj['Daytime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D77'] = day_net_obj['Daytime']['CNBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('E!'):
            dest_pp_sheet['E6'] = day_net_obj['Daytime']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E29'] = day_net_obj['Daytime']['E!']['target_index']
            dest_pp_sheet['E53'] = day_net_obj['Daytime']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E77'] = day_net_obj['Daytime']['E!']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Esquire'):
            dest_pp_sheet['F6'] = day_net_obj['Daytime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F29'] = day_net_obj['Daytime']['Esquire']['target_index']
            dest_pp_sheet['F53'] = day_net_obj['Daytime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F77'] = day_net_obj['Daytime']['Esquire']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Golf Channel'):
            dest_pp_sheet['G6'] = day_net_obj['Daytime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G29'] = day_net_obj['Daytime']['Golf Channel']['target_index']
            dest_pp_sheet['G53'] = day_net_obj['Daytime']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G77'] = day_net_obj['Daytime']['Golf Channel']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('MSNBC'):
            dest_pp_sheet['H6'] = day_net_obj['Daytime']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H29'] = day_net_obj['Daytime']['MSNBC']['target_index']
            dest_pp_sheet['H53'] = day_net_obj['Daytime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H77'] = day_net_obj['Daytime']['MSNBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('NBC'):
            dest_pp_sheet['I6'] = day_net_obj['Daytime']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I29'] = day_net_obj['Daytime']['NBC']['target_index']
            dest_pp_sheet['I53'] = day_net_obj['Daytime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I77'] = day_net_obj['Daytime']['NBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('NBCSN'):
            dest_pp_sheet['J6'] = day_net_obj['Daytime']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J29'] = day_net_obj['Daytime']['NBCSN']['target_index']
            dest_pp_sheet['J53'] = day_net_obj['Daytime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J77'] = day_net_obj['Daytime']['NBCSN']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Oxygen'):
            dest_pp_sheet['K6'] = day_net_obj['Daytime']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K29'] = day_net_obj['Daytime']['Oxygen']['target_index']
            dest_pp_sheet['K53'] = day_net_obj['Daytime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K77'] = day_net_obj['Daytime']['Oxygen']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Sprout'):
            dest_pp_sheet['L6'] = day_net_obj['Daytime']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L29'] = day_net_obj['Daytime']['Sprout']['target_index']
            dest_pp_sheet['L53'] = day_net_obj['Daytime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L77'] = day_net_obj['Daytime']['Sprout']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Syfy'):
            dest_pp_sheet['M6'] = day_net_obj['Daytime']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M29'] = day_net_obj['Daytime']['Syfy']['target_index']
            dest_pp_sheet['M53'] = day_net_obj['Daytime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M77'] = day_net_obj['Daytime']['Syfy']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Daytime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Telemundo'):
            dest_pp_sheet['N6'] = day_net_obj['Daytime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N29'] = day_net_obj['Daytime']['Telemundo']['target_index']
            dest_pp_sheet['N53'] = day_net_obj['Daytime']['Telemundo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['N77'] = day_net_obj['Daytime']['Telemundo']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Daytime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('USA'):
            dest_pp_sheet['O6'] = day_net_obj['Daytime']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O29'] = day_net_obj['Daytime']['USA']['target_index']
            dest_pp_sheet['O53'] = day_net_obj['Daytime']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O77'] = day_net_obj['Daytime']['USA']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Daytime']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Early Fringe
        fe_im_total = 0
        fe_re_total = 0
        if day_net_obj['Early Fringe'].has_key('Bravo'):
            dest_pp_sheet['B7'] = day_net_obj['Early Fringe']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B30'] = day_net_obj['Early Fringe']['Bravo']['target_index']
            dest_pp_sheet['B54'] = day_net_obj['Early Fringe']['Bravo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['B78'] = day_net_obj['Early Fringe']['Bravo']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Chiller'):
            dest_pp_sheet['C7'] = day_net_obj['Early Fringe']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C30'] = day_net_obj['Early Fringe']['Chiller']['target_index']
            dest_pp_sheet['C54'] = day_net_obj['Early Fringe']['Chiller']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['C78'] = day_net_obj['Early Fringe']['Chiller']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('CNBC'):
            dest_pp_sheet['D7'] = day_net_obj['Early Fringe']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D30'] = day_net_obj['Early Fringe']['CNBC']['target_index']
            dest_pp_sheet['D54'] = day_net_obj['Early Fringe']['CNBC']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['D78'] = day_net_obj['Early Fringe']['CNBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('E!'):
            dest_pp_sheet['E7'] = day_net_obj['Early Fringe']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E30'] = day_net_obj['Early Fringe']['E!']['target_index']
            dest_pp_sheet['E54'] = day_net_obj['Early Fringe']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E78'] = day_net_obj['Early Fringe']['E!']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Esquire'):
            dest_pp_sheet['F7'] = day_net_obj['Early Fringe']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F30'] = day_net_obj['Early Fringe']['Esquire']['target_index']
            dest_pp_sheet['F54'] = day_net_obj['Early Fringe']['Esquire']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['F78'] = day_net_obj['Early Fringe']['Esquire']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Golf Channel'):
            dest_pp_sheet['G7'] = day_net_obj['Early Fringe']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G30'] = day_net_obj['Early Fringe']['Golf Channel']['target_index']
            dest_pp_sheet['G54'] = day_net_obj['Early Fringe']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G78'] = day_net_obj['Early Fringe']['Golf Channel']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
        if day_net_obj['Early Fringe'].has_key('MSNBC'):
            dest_pp_sheet['H7'] = day_net_obj['Early Fringe']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H30'] = day_net_obj['Early Fringe']['MSNBC']['target_index']
            dest_pp_sheet['H54'] = day_net_obj['Early Fringe']['MSNBC']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['H78'] = day_net_obj['Early Fringe']['MSNBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('NBC'):
            dest_pp_sheet['I7'] = day_net_obj['Early Fringe']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I30'] = day_net_obj['Early Fringe']['NBC']['target_index']
            dest_pp_sheet['I54'] = day_net_obj['Early Fringe']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I78'] = day_net_obj['Early Fringe']['NBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('NBCSN'):
            dest_pp_sheet['J7'] = day_net_obj['Early Fringe']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J30'] = day_net_obj['Early Fringe']['NBCSN']['target_index']
            dest_pp_sheet['J54'] = day_net_obj['Early Fringe']['NBCSN']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['J78'] = day_net_obj['Early Fringe']['NBCSN']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Oxygen'):
            dest_pp_sheet['K7'] = day_net_obj['Early Fringe']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K30'] = day_net_obj['Early Fringe']['Oxygen']['target_index']
            dest_pp_sheet['K54'] = day_net_obj['Early Fringe']['Oxygen']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['K78'] = day_net_obj['Early Fringe']['Oxygen']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Sprout'):
            dest_pp_sheet['L7'] = day_net_obj['Early Fringe']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L30'] = day_net_obj['Early Fringe']['Sprout']['target_index']
            dest_pp_sheet['L54'] = day_net_obj['Early Fringe']['Sprout']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['L78'] = day_net_obj['Early Fringe']['Sprout']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Syfy'):
            dest_pp_sheet['M7'] = day_net_obj['Early Fringe']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M30'] = day_net_obj['Early Fringe']['Syfy']['target_index']
            dest_pp_sheet['M54'] = day_net_obj['Early Fringe']['Syfy']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['M78'] = day_net_obj['Early Fringe']['Syfy']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Telemundo'):
            dest_pp_sheet['N7'] = day_net_obj['Early Fringe']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N30'] = day_net_obj['Early Fringe']['Telemundo']['target_index']
            dest_pp_sheet['N54'] = day_net_obj['Early Fringe']['Telemundo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['N78'] = day_net_obj['Early Fringe']['Telemundo']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('USA'):
            dest_pp_sheet['O7'] = day_net_obj['Early Fringe']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O30'] = day_net_obj['Early Fringe']['USA']['target_index']
            dest_pp_sheet['O54'] = day_net_obj['Early Fringe']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O78'] = day_net_obj['Early Fringe']['USA']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Prime
        pr_im_total = 0
        pr_re_total = 0
        if day_net_obj['Prime'].has_key('Bravo'):
            dest_pp_sheet['B8'] = day_net_obj['Prime']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B31'] = day_net_obj['Prime']['Bravo']['target_index']
            dest_pp_sheet['B55'] = day_net_obj['Prime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B79'] = day_net_obj['Prime']['Bravo']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Chiller'):
            dest_pp_sheet['C8'] = day_net_obj['Prime']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C31'] = day_net_obj['Prime']['Chiller']['target_index']
            dest_pp_sheet['C55'] = day_net_obj['Prime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C79'] = day_net_obj['Prime']['Chiller']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('CNBC'):
            dest_pp_sheet['D8'] = day_net_obj['Prime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D31'] = day_net_obj['Prime']['CNBC']['target_index']
            dest_pp_sheet['D55'] = day_net_obj['Prime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D79'] = day_net_obj['Prime']['CNBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            pr_re_total += day_net_obj['Prime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('E!'):
            dest_pp_sheet['E8'] = day_net_obj['Prime']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E31'] = day_net_obj['Prime']['E!']['target_index']
            dest_pp_sheet['E55'] = day_net_obj['Prime']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E79'] = day_net_obj['Prime']['E!']['target_frequency']
            pr_im_total += day_net_obj['Prime']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Esquire'):
            dest_pp_sheet['F8'] = day_net_obj['Prime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F31'] = day_net_obj['Prime']['Esquire']['target_index']
            dest_pp_sheet['F55'] = day_net_obj['Prime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F79'] = day_net_obj['Prime']['Esquire']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            pr_re_total += day_net_obj['Prime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Golf Channel'):
            dest_pp_sheet['G8'] = day_net_obj['Prime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G31'] = day_net_obj['Prime']['Golf Channel']['target_index']
            dest_pp_sheet['G55'] = day_net_obj['Prime']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G79'] = day_net_obj['Prime']['Golf Channel']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            pr_re_total += day_net_obj['Prime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('MSNBC'):
            dest_pp_sheet['H8'] = day_net_obj['Prime']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H31'] = day_net_obj['Prime']['MSNBC']['target_index']
            dest_pp_sheet['H55'] = day_net_obj['Prime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H79'] = day_net_obj['Prime']['MSNBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('NBC'):
            dest_pp_sheet['I8'] = day_net_obj['Prime']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I31'] = day_net_obj['Prime']['NBC']['target_index']
            dest_pp_sheet['I55'] = day_net_obj['Prime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I79'] = day_net_obj['Prime']['NBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('NBCSN'):
            dest_pp_sheet['J8'] = day_net_obj['Prime']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J31'] = day_net_obj['Prime']['NBCSN']['target_index']
            dest_pp_sheet['J55'] = day_net_obj['Prime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J79'] = day_net_obj['Prime']['NBCSN']['target_frequency']
            pr_im_total += day_net_obj['Prime']['NBCSN']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Oxygen'):
            dest_pp_sheet['K8'] = day_net_obj['Prime']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K31'] = day_net_obj['Prime']['Oxygen']['target_index']
            dest_pp_sheet['K55'] = day_net_obj['Prime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K79'] = day_net_obj['Prime']['Oxygen']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Sprout'):
            dest_pp_sheet['L8'] = day_net_obj['Prime']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L31'] = day_net_obj['Prime']['Sprout']['target_index']
            dest_pp_sheet['L55'] = day_net_obj['Prime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L79'] = day_net_obj['Prime']['Sprout']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Syfy'):
            dest_pp_sheet['M8'] = day_net_obj['Prime']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M31'] = day_net_obj['Prime']['Syfy']['target_index']
            dest_pp_sheet['M55'] = day_net_obj['Prime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M79'] = day_net_obj['Prime']['Syfy']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Telemundo'):
            dest_pp_sheet['N8'] = day_net_obj['Prime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N31'] = day_net_obj['Prime']['Telemundo']['target_index']
            dest_pp_sheet['N55'] = day_net_obj['Prime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N79'] = day_net_obj['Prime']['Telemundo']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            pr_re_total += day_net_obj['Prime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('USA'):
            dest_pp_sheet['O8'] = day_net_obj['Prime']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O31'] = day_net_obj['Prime']['USA']['target_index']
            dest_pp_sheet['O55'] = day_net_obj['Prime']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O79'] = day_net_obj['Prime']['USA']['target_frequency']
            pr_im_total += day_net_obj['Prime']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Late Night
        ln_im_total = 0
        ln_re_total = 0
        if day_net_obj['Late Night'].has_key('Bravo'):
            dest_pp_sheet['B9'] = day_net_obj['Late Night']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B32'] = day_net_obj['Late Night']['Bravo']['target_index']
            dest_pp_sheet['B56'] = day_net_obj['Late Night']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B80'] = day_net_obj['Late Night']['Bravo']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Chiller'):
            dest_pp_sheet['C9'] = day_net_obj['Late Night']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C32'] = day_net_obj['Late Night']['Chiller']['target_index']
            dest_pp_sheet['C56'] = day_net_obj['Late Night']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C80'] = day_net_obj['Late Night']['Chiller']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('CNBC'):
            dest_pp_sheet['D9'] = day_net_obj['Late Night']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D32'] = day_net_obj['Late Night']['CNBC']['target_index']
            dest_pp_sheet['D56'] = day_net_obj['Late Night']['CNBC']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['D80'] = day_net_obj['Late Night']['CNBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('E!'):
            dest_pp_sheet['E9'] = day_net_obj['Late Night']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E32'] = day_net_obj['Late Night']['E!']['target_index']
            dest_pp_sheet['E56'] = day_net_obj['Late Night']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E80'] = day_net_obj['Late Night']['E!']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Esquire'):
            dest_pp_sheet['F9'] = day_net_obj['Late Night']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F32'] = day_net_obj['Late Night']['Esquire']['target_index']
            dest_pp_sheet['F56'] = day_net_obj['Late Night']['Esquire']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['F80'] = day_net_obj['Late Night']['Esquire']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Golf Channel'):
            dest_pp_sheet['G9'] = day_net_obj['Late Night']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G32'] = day_net_obj['Late Night']['Golf Channel']['target_index']
            dest_pp_sheet['G56'] = day_net_obj['Late Night']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G80'] = day_net_obj['Late Night']['Golf Channel']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('MSNBC'):
            dest_pp_sheet['H9'] = day_net_obj['Late Night']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H32'] = day_net_obj['Late Night']['MSNBC']['target_index']
            dest_pp_sheet['H56'] = day_net_obj['Late Night']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H80'] = day_net_obj['Late Night']['MSNBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('NBC'):
            dest_pp_sheet['I9'] = day_net_obj['Late Night']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I32'] = day_net_obj['Late Night']['NBC']['target_index']
            dest_pp_sheet['I56'] = day_net_obj['Late Night']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I80'] = day_net_obj['Late Night']['NBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('NBCSN'):
            dest_pp_sheet['J9'] = day_net_obj['Late Night']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J32'] = day_net_obj['Late Night']['NBCSN']['target_index']
            dest_pp_sheet['J56'] = day_net_obj['Late Night']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J80'] = day_net_obj['Late Night']['NBCSN']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Oxygen'):
            dest_pp_sheet['K9'] = day_net_obj['Late Night']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K32'] = day_net_obj['Late Night']['Oxygen']['target_index']
            dest_pp_sheet['K56'] = day_net_obj['Late Night']['Oxygen']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['K80'] = day_net_obj['Late Night']['Oxygen']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Sprout'):
            dest_pp_sheet['L9'] = day_net_obj['Late Night']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L32'] = day_net_obj['Late Night']['Sprout']['target_index']
            dest_pp_sheet['L56'] = day_net_obj['Late Night']['Sprout']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['L80'] = day_net_obj['Late Night']['Sprout']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Syfy'):
            dest_pp_sheet['M9'] = day_net_obj['Late Night']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M32'] = day_net_obj['Late Night']['Syfy']['target_index']
            dest_pp_sheet['M56'] = day_net_obj['Late Night']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M80'] = day_net_obj['Late Night']['Syfy']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Telemundo'):
            dest_pp_sheet['N9'] = day_net_obj['Late Night']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N32'] = day_net_obj['Late Night']['Telemundo']['target_index']
            dest_pp_sheet['N56'] = day_net_obj['Late Night']['Telemundo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['N80'] = day_net_obj['Late Night']['Telemundo']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('USA'):
            dest_pp_sheet['O9'] = day_net_obj['Late Night']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O32'] = day_net_obj['Late Night']['USA']['target_index']
            dest_pp_sheet['O56'] = day_net_obj['Late Night']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O80'] = day_net_obj['Late Night']['USA']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Overnight
        on_im_total = 0
        on_re_total = 0
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Bravo'):
            dest_pp_sheet['B10'] = day_net_obj['Overnight']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B33'] = day_net_obj['Overnight']['Bravo']['target_index']
            dest_pp_sheet['B57'] = day_net_obj['Overnight']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B81'] = day_net_obj['Overnight']['Bravo']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Chiller'):
            dest_pp_sheet['C10'] = day_net_obj['Overnight']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C33'] = day_net_obj['Overnight']['Chiller']['target_index']
            dest_pp_sheet['C57'] = day_net_obj['Overnight']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C81'] = day_net_obj['Overnight']['Chiller']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('CNBC'):
            dest_pp_sheet['D10'] = day_net_obj['Overnight']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D33'] = day_net_obj['Overnight']['CNBC']['target_index']
            dest_pp_sheet['D57'] = day_net_obj['Overnight']['CNBC']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['D81'] = day_net_obj['Overnight']['CNBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('E!'):
            dest_pp_sheet['E10'] = day_net_obj['Overnight']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E33'] = day_net_obj['Overnight']['E!']['target_index']
            dest_pp_sheet['E57'] = day_net_obj['Overnight']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E81'] = day_net_obj['Overnight']['E!']['target_frequency']
            on_im_total += day_net_obj['Overnight']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Esquire'):
            dest_pp_sheet['F10'] = day_net_obj['Overnight']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F33'] = day_net_obj['Overnight']['Esquire']['target_index']
            dest_pp_sheet['F57'] = day_net_obj['Overnight']['Esquire']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['F81'] = day_net_obj['Overnight']['Esquire']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Golf Channel'):
            dest_pp_sheet['G10'] = day_net_obj['Overnight']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G33'] = day_net_obj['Overnight']['Golf Channel']['target_index']
            dest_pp_sheet['G57'] = day_net_obj['Overnight']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G81'] = day_net_obj['Overnight']['Golf Channel']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('MSNBC'):
            dest_pp_sheet['H10'] = day_net_obj['Overnight']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H33'] = day_net_obj['Overnight']['MSNBC']['target_index']
            dest_pp_sheet['H57'] = day_net_obj['Overnight']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H81'] = day_net_obj['Overnight']['MSNBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('NBC'):
            dest_pp_sheet['I10'] = day_net_obj['Overnight']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I33'] = day_net_obj['Overnight']['NBC']['target_index']
            dest_pp_sheet['I57'] = day_net_obj['Overnight']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I81'] = day_net_obj['Overnight']['NBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('NBCSN'):
            dest_pp_sheet['J10'] = day_net_obj['Overnight']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J33'] = day_net_obj['Overnight']['NBCSN']['target_index']
            dest_pp_sheet['J57'] = day_net_obj['Overnight']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J81'] = day_net_obj['Overnight']['NBCSN']['target_frequency']
            on_im_total += day_net_obj['Overnight']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Oxygen'):
            dest_pp_sheet['K10'] = day_net_obj['Overnight']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K33'] = day_net_obj['Overnight']['Oxygen']['target_index']
            dest_pp_sheet['K57'] = day_net_obj['Overnight']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K81'] = day_net_obj['Overnight']['Oxygen']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Sprout'):
            dest_pp_sheet['L10'] = day_net_obj['Overnight']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L33'] = day_net_obj['Overnight']['Sprout']['target_index']
            dest_pp_sheet['L57'] = day_net_obj['Overnight']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L81'] = day_net_obj['Overnight']['Sprout']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Syfy'):
            dest_pp_sheet['M10'] = day_net_obj['Overnight']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M33'] = day_net_obj['Overnight']['Syfy']['target_index']
            dest_pp_sheet['M57'] = day_net_obj['Overnight']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M81'] = day_net_obj['Overnight']['Syfy']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Telemundo'):
            dest_pp_sheet['N10'] = day_net_obj['Overnight']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N33'] = day_net_obj['Overnight']['Telemundo']['target_index']
            dest_pp_sheet['N57'] = day_net_obj['Overnight']['Telemundo']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['N81'] = day_net_obj['Overnight']['Telemundo']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('USA'):
            dest_pp_sheet['O10'] = day_net_obj['Overnight']['USA']['target_impressions'] / net_obj['Total'][
                'target_impOessions']
            dest_pp_sheet['O33'] = day_net_obj['Overnight']['USA']['target_index']
            dest_pp_sheet['O57'] = day_net_obj['Overnight']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O81'] = day_net_obj['Overnight']['USA']['target_frequency']
            on_im_total += day_net_obj['Overnight']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['USA']['target_reach'] / net_obj['Total']['target_reach']

        dest_pp_sheet['P5'] = mo_im_total
        dest_pp_sheet['P6'] = dy_im_total
        dest_pp_sheet['P7'] = fe_im_total
        dest_pp_sheet['P8'] = pr_im_total
        dest_pp_sheet['P9'] = ln_im_total
        dest_pp_sheet['P10'] = on_im_total

        dest_pp_sheet['P52'] = mo_re_total
        dest_pp_sheet['P53'] = dy_re_total
        dest_pp_sheet['P54'] = fe_re_total
        dest_pp_sheet['P55'] = pr_re_total
        dest_pp_sheet['P56'] = ln_re_total
        dest_pp_sheet['P57'] = on_re_total

        dest_pp_sheet['P28'] = day_net_obj['Morning']['Total']['target_index'] if day_net_obj['Morning'].has_key(
            'Total') else ''
        dest_pp_sheet['P29'] = day_net_obj['Daytime']['Total']['target_index'] if day_net_obj['Daytime'].has_key(
            'Total') else ''
        dest_pp_sheet['P30'] = day_net_obj['Early Fringe']['Total']['target_index'] if day_net_obj[
            'Early Fringe'].has_key(
            'Total') else ''
        dest_pp_sheet['P31'] = day_net_obj['Prime']['Total']['target_index'] if day_net_obj['Prime'].has_key(
            'Total') else ''
        dest_pp_sheet['P32'] = day_net_obj['Late Night']['Total']['target_index'] if day_net_obj['Late Night'].has_key(
            'Total') else ''
        dest_pp_sheet['P33'] = day_net_obj['Overnight']['Total']['target_index'] if day_net_obj['Overnight'].has_key(
            'Total') else ''

        dest_pp_sheet['P76'] = day_net_obj['Morning']['Total']['target_frequency'] if day_net_obj['Morning'].has_key(
            'Total') else ''
        dest_pp_sheet['P77'] = day_net_obj['Daytime']['Total']['target_frequency'] if day_net_obj['Daytime'].has_key(
            'Total') else ''
        dest_pp_sheet['P78'] = day_net_obj['Early Fringe']['Total']['target_frequency'] if day_net_obj[
            'Early Fringe'].has_key('Total') else ''
        dest_pp_sheet['P79'] = day_net_obj['Prime']['Total']['target_frequency'] if day_net_obj['Prime'].has_key(
            'Total') else ''
        dest_pp_sheet['P80'] = day_net_obj['Late Night']['Total']['target_frequency'] if day_net_obj[
            'Late Night'].has_key(
            'Total') else ''
        dest_pp_sheet['P81'] = day_net_obj['Overnight']['Total']['target_frequency'] if day_net_obj[
            'Overnight'].has_key(
            'Total') else ''

        # Totals
        dest_pp_sheet['B11'] = net_obj['Bravo']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C11'] = net_obj['Chiller']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D11'] = net_obj['CNBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E11'] = net_obj['E!']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F11'] = net_obj['Esquire']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G11'] = net_obj['Golf Channel']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H11'] = net_obj['MSNBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I11'] = net_obj['NBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J11'] = net_obj['NBCSN']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K11'] = net_obj['Oxygen']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L11'] = net_obj['Sprout']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M11'] = net_obj['Syfy']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N11'] = net_obj['Telemundo']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['O11'] = net_obj['USA']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['P11'] = net_obj['Total']['target_impressions'] / net_obj['Total']['target_impressions']

        dest_pp_sheet['B34'] = net_obj['Bravo']['target_index'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C34'] = net_obj['Chiller']['target_index'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D34'] = net_obj['CNBC']['target_index'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E34'] = net_obj['E!']['target_index'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F34'] = net_obj['Esquire']['target_index'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G34'] = net_obj['Golf Channel']['target_index'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H34'] = net_obj['MSNBC']['target_index'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I34'] = net_obj['NBC']['target_index'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J34'] = net_obj['NBCSN']['target_index'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K34'] = net_obj['Oxygen']['target_index'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L34'] = net_obj['Sprout']['target_index'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M34'] = net_obj['Syfy']['target_index'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N34'] = net_obj['Telemundo']['target_index'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['O34'] = net_obj['USA']['target_index'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['P34'] = net_obj['Total']['target_index']

        dest_pp_sheet['B58'] = net_obj['Bravo']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Bravo') else ''
        dest_pp_sheet['C58'] = net_obj['Chiller']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Chiller') else ''
        dest_pp_sheet['D58'] = net_obj['CNBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'CNBC') else ''
        dest_pp_sheet['E58'] = net_obj['E!']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'E!') else ''
        dest_pp_sheet['F58'] = net_obj['Esquire']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Esquire') else ''
        dest_pp_sheet['G58'] = net_obj['Golf Channel']['target_reach'] / net_obj['Total'][
            'target_reach'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H58'] = net_obj['MSNBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'MSNBC') else ''
        dest_pp_sheet['I58'] = net_obj['NBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'NBC') else ''
        dest_pp_sheet['J58'] = net_obj['NBCSN']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'NBCSN') else ''
        dest_pp_sheet['K58'] = net_obj['Oxygen']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Oxygen') else ''
        dest_pp_sheet['L58'] = net_obj['Sprout']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Sprout') else ''
        dest_pp_sheet['M58'] = net_obj['Syfy']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Syfy') else ''
        dest_pp_sheet['N58'] = net_obj['Telemundo']['target_reach'] / net_obj['Total'][
            'target_reach'] if net_obj.has_key(
            'Telemundo') else ''
        dest_pp_sheet['O58'] = net_obj['USA']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'USA') else ''
        dest_pp_sheet['P58'] = net_obj['Total']['target_reach'] / net_obj['Total']['target_reach']

        dest_pp_sheet['B82'] = net_obj['Bravo']['target_frequency'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C82'] = net_obj['Chiller']['target_frequency'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D82'] = net_obj['CNBC']['target_frequency'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E82'] = net_obj['E!']['target_frequency'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F82'] = net_obj['Esquire']['target_frequency'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G82'] = net_obj['Golf Channel']['target_frequency'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H82'] = net_obj['MSNBC']['target_frequency'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I82'] = net_obj['NBC']['target_frequency'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J82'] = net_obj['NBCSN']['target_frequency'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K82'] = net_obj['Oxygen']['target_frequency'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L82'] = net_obj['Sprout']['target_frequency'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M82'] = net_obj['Syfy']['target_frequency'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N82'] = net_obj['Telemundo']['target_frequency'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['N82'] = net_obj['USA']['target_frequency'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['O82'] = net_obj['Total']['target_frequency']

        dest_pp_sheet['B100'] = net_obj['Bravo']['target_frequency'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['B101'] = net_obj['Chiller']['target_frequency'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['B102'] = net_obj['CNBC']['target_frequency'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['B103'] = net_obj['E!']['target_frequency'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['B104'] = net_obj['Esquire']['target_frequency'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['B105'] = net_obj['Golf Channel']['target_frequency'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['B106'] = net_obj['MSNBC']['target_frequency'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['B107'] = net_obj['NBC']['target_frequency'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['B108'] = net_obj['NBCSN']['target_frequency'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['B109'] = net_obj['Oxygen']['target_frequency'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['B110'] = net_obj['Sprout']['target_frequency'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['B111'] = net_obj['Syfy']['target_frequency'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['B112'] = net_obj['Telemundo']['target_frequency'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['B113'] = net_obj['USA']['target_frequency'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['B99'] = net_obj['Total']['target_frequency']

        dest_pp_sheet['B117'] = day_net_obj['Morning']['NBC']['target_frequency'] if day_net_obj['Morning'].has_key(
            'NBC') else ''
        dest_pp_sheet['B118'] = day_net_obj['Daytime']['NBC']['target_frequency'] if day_net_obj['Daytime'].has_key(
            'NBC') else ''
        dest_pp_sheet['B119'] = day_net_obj['Early Fringe']['NBC']['target_frequency'] if day_net_obj[
            'Early Fringe'].has_key('NBC') else ''
        dest_pp_sheet['B120'] = day_net_obj['Prime']['NBC']['target_frequency'] if day_net_obj['Prime'].has_key(
            'NBC') else ''
        dest_pp_sheet['B121'] = day_net_obj['Late Night']['NBC']['target_frequency'] if day_net_obj[
            'Late Night'].has_key(
            'NBC') else ''
        dest_pp_sheet['B122'] = day_net_obj['Overnight']['NBC']['target_frequency'] if day_net_obj.has_key(
            'Overnight') and \
                                                                                       day_net_obj['Overnight'].has_key(
                                                                                           'NBC') else ''
    else:
        # Morning
        mo_im_total = 0
        mo_re_total = 0
        if day_net_obj['Morning'].has_key('Bravo'):
            dest_pp_sheet['B16'] = day_net_obj['Morning']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B39'] = day_net_obj['Morning']['Bravo']['target_index']
            dest_pp_sheet['B63'] = day_net_obj['Morning']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B88'] = day_net_obj['Morning']['Bravo']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Chiller'):
            dest_pp_sheet['C16'] = day_net_obj['Morning']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C39'] = day_net_obj['Morning']['Chiller']['target_index']
            dest_pp_sheet['C63'] = day_net_obj['Morning']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C88'] = day_net_obj['Morning']['Chiller']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('CNBC'):
            dest_pp_sheet['D16'] = day_net_obj['Morning']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D39'] = day_net_obj['Morning']['CNBC']['target_index']
            dest_pp_sheet['D63'] = day_net_obj['Morning']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D88'] = day_net_obj['Morning']['CNBC']['target_frequency']
            mo_im_total += day_net_obj['Morning']['CNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('E!'):
            dest_pp_sheet['E16'] = day_net_obj['Morning']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E39'] = day_net_obj['Morning']['E!']['target_index']
            dest_pp_sheet['E63'] = day_net_obj['Morning']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E88'] = day_net_obj['Morning']['E!']['target_frequency']
            mo_im_total += day_net_obj['Morning']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Esquire'):
            dest_pp_sheet['F16'] = day_net_obj['Morning']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F39'] = day_net_obj['Morning']['Esquire']['target_index']
            dest_pp_sheet['F63'] = day_net_obj['Morning']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F88'] = day_net_obj['Morning']['Esquire']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Esquire']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Golf Channel'):
            dest_pp_sheet['G16'] = day_net_obj['Morning']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G39'] = day_net_obj['Morning']['Golf Channel']['target_index']
            dest_pp_sheet['G63'] = day_net_obj['Morning']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['G88'] = day_net_obj['Morning']['Golf Channel']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            mo_re_total += day_net_obj['Morning']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('MSNBC'):
            dest_pp_sheet['H16'] = day_net_obj['Morning']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H39'] = day_net_obj['Morning']['MSNBC']['target_index']
            dest_pp_sheet['H63'] = day_net_obj['Morning']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            mo_im_total += day_net_obj['Morning']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('NBC'):
            dest_pp_sheet['I16'] = day_net_obj['Morning']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I39'] = day_net_obj['Morning']['NBC']['target_index']
            dest_pp_sheet['I63'] = day_net_obj['Morning']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I88'] = day_net_obj['Morning']['NBC']['target_frequency']
            mo_im_total += day_net_obj['Morning']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('NBCSN'):
            dest_pp_sheet['J16'] = day_net_obj['Morning']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J39'] = day_net_obj['Morning']['NBCSN']['target_index']
            dest_pp_sheet['J63'] = day_net_obj['Morning']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J88'] = day_net_obj['Morning']['NBCSN']['target_frequency']
            mo_im_total += day_net_obj['Morning']['NBCSN']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Oxygen'):
            dest_pp_sheet['K16'] = day_net_obj['Morning']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K39'] = day_net_obj['Morning']['Oxygen']['target_index']
            dest_pp_sheet['K63'] = day_net_obj['Morning']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K88'] = day_net_obj['Morning']['Oxygen']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Sprout'):
            dest_pp_sheet['L16'] = day_net_obj['Morning']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L39'] = day_net_obj['Morning']['Sprout']['target_index']
            dest_pp_sheet['L63'] = day_net_obj['Morning']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L88'] = day_net_obj['Morning']['Sprout']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Syfy'):
            dest_pp_sheet['M16'] = day_net_obj['Morning']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M39'] = day_net_obj['Morning']['Syfy']['target_index']
            dest_pp_sheet['M63'] = day_net_obj['Morning']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M88'] = day_net_obj['Morning']['Syfy']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('Telemundo'):
            dest_pp_sheet['N16'] = day_net_obj['Morning']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N39'] = day_net_obj['Morning']['Telemundo']['target_index']
            dest_pp_sheet['N63'] = day_net_obj['Morning']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N88'] = day_net_obj['Morning']['Telemundo']['target_frequency']
            mo_im_total += day_net_obj['Morning']['Telemundo']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Morning'].has_key('USA'):
            dest_pp_sheet['O16'] = day_net_obj['Morning']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O39'] = day_net_obj['Morning']['USA']['target_index']
            dest_pp_sheet['O63'] = day_net_obj['Morning']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O88'] = day_net_obj['Morning']['USA']['target_frequency']
            mo_im_total += day_net_obj['Morning']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Morning']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Daytime
        dy_im_total = 0
        dy_re_total = 0
        if day_net_obj['Daytime'].has_key('Bravo'):
            dest_pp_sheet['B17'] = day_net_obj['Daytime']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B40'] = day_net_obj['Daytime']['Bravo']['target_index']
            dest_pp_sheet['B64'] = day_net_obj['Daytime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B89'] = day_net_obj['Daytime']['Bravo']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Chiller'):
            dest_pp_sheet['C17'] = day_net_obj['Daytime']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C40'] = day_net_obj['Daytime']['Chiller']['target_index']
            dest_pp_sheet['C64'] = day_net_obj['Daytime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C89'] = day_net_obj['Daytime']['Chiller']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('CNBC'):
            dest_pp_sheet['D17'] = day_net_obj['Daytime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D40'] = day_net_obj['Daytime']['CNBC']['target_index']
            dest_pp_sheet['D64'] = day_net_obj['Daytime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D89'] = day_net_obj['Daytime']['CNBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['CNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('E!'):
            dest_pp_sheet['E17'] = day_net_obj['Daytime']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E40'] = day_net_obj['Daytime']['E!']['target_index']
            dest_pp_sheet['E64'] = day_net_obj['Daytime']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E89'] = day_net_obj['Daytime']['E!']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Esquire'):
            dest_pp_sheet['F17'] = day_net_obj['Daytime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F40'] = day_net_obj['Daytime']['Esquire']['target_index']
            dest_pp_sheet['F64'] = day_net_obj['Daytime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F89'] = day_net_obj['Daytime']['Esquire']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Esquire']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Golf Channel'):
            dest_pp_sheet['G17'] = day_net_obj['Daytime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G40'] = day_net_obj['Daytime']['Golf Channel']['target_index']
            dest_pp_sheet['G64'] = day_net_obj['Daytime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['G89'] = day_net_obj['Daytime']['Golf Channel']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dy_re_total += day_net_obj['Daytime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('MSNBC'):
            dest_pp_sheet['H17'] = day_net_obj['Daytime']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H40'] = day_net_obj['Daytime']['MSNBC']['target_index']
            dest_pp_sheet['H64'] = day_net_obj['Daytime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H89'] = day_net_obj['Daytime']['MSNBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('NBC'):
            dest_pp_sheet['I17'] = day_net_obj['Daytime']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I40'] = day_net_obj['Daytime']['NBC']['target_index']
            dest_pp_sheet['I64'] = day_net_obj['Daytime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I89'] = day_net_obj['Daytime']['NBC']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('NBCSN'):
            dest_pp_sheet['J17'] = day_net_obj['Daytime']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J40'] = day_net_obj['Daytime']['NBCSN']['target_index']
            dest_pp_sheet['J64'] = day_net_obj['Daytime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J89'] = day_net_obj['Daytime']['NBCSN']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['NBCSN']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Oxygen'):
            dest_pp_sheet['K17'] = day_net_obj['Daytime']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K40'] = day_net_obj['Daytime']['Oxygen']['target_index']
            dest_pp_sheet['K64'] = day_net_obj['Daytime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K89'] = day_net_obj['Daytime']['Oxygen']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Sprout'):
            dest_pp_sheet['L17'] = day_net_obj['Daytime']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L40'] = day_net_obj['Daytime']['Sprout']['target_index']
            dest_pp_sheet['L64'] = day_net_obj['Daytime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L89'] = day_net_obj['Daytime']['Sprout']['target_frequency']
            dy_im_total += day_net_obj['Daytime']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            dy_re_total += day_net_obj['Daytime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Syfy'):
            dest_pp_sheet['M17'] = day_net_obj['Daytime']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M40'] = day_net_obj['Daytime']['Syfy']['target_index']
            dest_pp_sheet['M64'] = day_net_obj['Daytime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M89'] = day_net_obj['Daytime']['Syfy']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Daytime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('Telemundo'):
            dest_pp_sheet['N17'] = day_net_obj['Daytime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N40'] = day_net_obj['Daytime']['Telemundo']['target_index']
            dest_pp_sheet['N64'] = day_net_obj['Daytime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N89'] = day_net_obj['Daytime']['Telemundo']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['Telemundo']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Daytime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Daytime'].has_key('USA'):
            dest_pp_sheet['O17'] = day_net_obj['Daytime']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O40'] = day_net_obj['Daytime']['USA']['target_index']
            dest_pp_sheet['O64'] = day_net_obj['Daytime']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O89'] = day_net_obj['Daytime']['USA']['target_frequency']
            mo_im_total += day_net_obj['Daytime']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            mo_re_total += day_net_obj['Daytime']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Early Fringe
        fe_im_total = 0
        fe_re_total = 0
        if day_net_obj['Early Fringe'].has_key('Bravo'):
            dest_pp_sheet['B18'] = day_net_obj['Early Fringe']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B41'] = day_net_obj['Early Fringe']['Bravo']['target_index']
            dest_pp_sheet['B65'] = day_net_obj['Early Fringe']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B90'] = day_net_obj['Early Fringe']['Bravo']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Chiller'):
            dest_pp_sheet['C18'] = day_net_obj['Early Fringe']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C41'] = day_net_obj['Early Fringe']['Chiller']['target_index']
            dest_pp_sheet['C65'] = day_net_obj['Early Fringe']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C90'] = day_net_obj['Early Fringe']['Chiller']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('CNBC'):
            dest_pp_sheet['D18'] = day_net_obj['Early Fringe']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D41'] = day_net_obj['Early Fringe']['CNBC']['target_index']
            dest_pp_sheet['D65'] = day_net_obj['Early Fringe']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D90'] = day_net_obj['Early Fringe']['CNBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('E!'):
            dest_pp_sheet['E18'] = day_net_obj['Early Fringe']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E41'] = day_net_obj['Early Fringe']['E!']['target_index']
            dest_pp_sheet['E65'] = day_net_obj['Early Fringe']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E90'] = day_net_obj['Early Fringe']['E!']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Esquire'):
            dest_pp_sheet['F18'] = day_net_obj['Early Fringe']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F41'] = day_net_obj['Early Fringe']['Esquire']['target_index']
            dest_pp_sheet['F65'] = day_net_obj['Early Fringe']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F90'] = day_net_obj['Early Fringe']['Esquire']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Golf Channel'):
            dest_pp_sheet['G18'] = day_net_obj['Early Fringe']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G41'] = day_net_obj['Early Fringe']['Golf Channel']['target_index']
            dest_pp_sheet['G65'] = day_net_obj['Early Fringe']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G90'] = day_net_obj['Early Fringe']['Golf Channel']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('MSNBC'):
            dest_pp_sheet['H18'] = day_net_obj['Early Fringe']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H41'] = day_net_obj['Early Fringe']['MSNBC']['target_index']
            dest_pp_sheet['H65'] = day_net_obj['Early Fringe']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H90'] = day_net_obj['Early Fringe']['MSNBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('NBC'):
            dest_pp_sheet['I18'] = day_net_obj['Early Fringe']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I41'] = day_net_obj['Early Fringe']['NBC']['target_index']
            dest_pp_sheet['I65'] = day_net_obj['Early Fringe']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I90'] = day_net_obj['Early Fringe']['NBC']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('NBCSN'):
            dest_pp_sheet['J18'] = day_net_obj['Early Fringe']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J41'] = day_net_obj['Early Fringe']['NBCSN']['target_index']
            dest_pp_sheet['J65'] = day_net_obj['Early Fringe']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J90'] = day_net_obj['Early Fringe']['NBCSN']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Oxygen'):
            dest_pp_sheet['K18'] = day_net_obj['Early Fringe']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K41'] = day_net_obj['Early Fringe']['Oxygen']['target_index']
            dest_pp_sheet['K65'] = day_net_obj['Early Fringe']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K90'] = day_net_obj['Early Fringe']['Oxygen']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Sprout'):
            dest_pp_sheet['L18'] = day_net_obj['Early Fringe']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L41'] = day_net_obj['Early Fringe']['Sprout']['target_index']
            dest_pp_sheet['L65'] = day_net_obj['Early Fringe']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L90'] = day_net_obj['Early Fringe']['Sprout']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Syfy'):
            dest_pp_sheet['M18'] = day_net_obj['Early Fringe']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M41'] = day_net_obj['Early Fringe']['Syfy']['target_index']
            dest_pp_sheet['M65'] = day_net_obj['Early Fringe']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M90'] = day_net_obj['Early Fringe']['Syfy']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('Telemundo'):
            dest_pp_sheet['N18'] = day_net_obj['Early Fringe']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N41'] = day_net_obj['Early Fringe']['Telemundo']['target_index']
            dest_pp_sheet['N65'] = day_net_obj['Early Fringe']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N90'] = day_net_obj['Early Fringe']['Telemundo']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Early Fringe'].has_key('USA'):
            dest_pp_sheet['O18'] = day_net_obj['Early Fringe']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O41'] = day_net_obj['Early Fringe']['USA']['target_index']
            dest_pp_sheet['O65'] = day_net_obj['Early Fringe']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O90'] = day_net_obj['Early Fringe']['USA']['target_frequency']
            fe_im_total += day_net_obj['Early Fringe']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            fe_re_total += day_net_obj['Early Fringe']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Prime
        pr_im_total = 0
        pr_re_total = 0
        if day_net_obj['Prime'].has_key('Bravo'):
            dest_pp_sheet['B19'] = day_net_obj['Prime']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B42'] = day_net_obj['Prime']['Bravo']['target_index']
            dest_pp_sheet['B66'] = day_net_obj['Prime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B91'] = day_net_obj['Prime']['Bravo']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Chiller'):
            dest_pp_sheet['C19'] = day_net_obj['Prime']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C42'] = day_net_obj['Prime']['Chiller']['target_index']
            dest_pp_sheet['C66'] = day_net_obj['Prime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C91'] = day_net_obj['Prime']['Chiller']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('CNBC'):
            dest_pp_sheet['D19'] = day_net_obj['Prime']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D42'] = day_net_obj['Prime']['CNBC']['target_index']
            dest_pp_sheet['D66'] = day_net_obj['Prime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D91'] = day_net_obj['Prime']['CNBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['CNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('E!'):
            dest_pp_sheet['E19'] = day_net_obj['Prime']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            dest_pp_sheet['E42'] = day_net_obj['Prime']['E!']['target_index']
            dest_pp_sheet['E66'] = day_net_obj['Prime']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E91'] = day_net_obj['Prime']['E!']['target_frequency']
            pr_im_total += day_net_obj['Prime']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Esquire'):
            dest_pp_sheet['F19'] = day_net_obj['Prime']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F42'] = day_net_obj['Prime']['Esquire']['target_index']
            dest_pp_sheet['F66'] = day_net_obj['Prime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F91'] = day_net_obj['Prime']['Esquire']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Esquire']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Golf Channel'):
            dest_pp_sheet['G19'] = day_net_obj['Prime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G42'] = day_net_obj['Prime']['Golf Channel']['target_index']
            dest_pp_sheet['G66'] = day_net_obj['Prime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['G91'] = day_net_obj['Prime']['Golf Channel']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            pr_re_total += day_net_obj['Prime']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('MSNBC'):
            dest_pp_sheet['H19'] = day_net_obj['Prime']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H42'] = day_net_obj['Prime']['MSNBC']['target_index']
            dest_pp_sheet['H66'] = day_net_obj['Prime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H91'] = day_net_obj['Prime']['MSNBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('NBC'):
            dest_pp_sheet['I19'] = day_net_obj['Prime']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I42'] = day_net_obj['Prime']['NBC']['target_index']
            dest_pp_sheet['I66'] = day_net_obj['Prime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I91'] = day_net_obj['Prime']['NBC']['target_frequency']
            pr_im_total += day_net_obj['Prime']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('NBCSN'):
            dest_pp_sheet['J19'] = day_net_obj['Prime']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J42'] = day_net_obj['Prime']['NBCSN']['target_index']
            dest_pp_sheet['J66'] = day_net_obj['Prime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J91'] = day_net_obj['Prime']['NBCSN']['target_frequency']
            pr_im_total += day_net_obj['Prime']['NBCSN']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Oxygen'):
            dest_pp_sheet['K19'] = day_net_obj['Prime']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K42'] = day_net_obj['Prime']['Oxygen']['target_index']
            dest_pp_sheet['K66'] = day_net_obj['Prime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K91'] = day_net_obj['Prime']['Oxygen']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Sprout'):
            dest_pp_sheet['L19'] = day_net_obj['Prime']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L42'] = day_net_obj['Prime']['Sprout']['target_index']
            dest_pp_sheet['L66'] = day_net_obj['Prime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L91'] = day_net_obj['Prime']['Sprout']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Syfy'):
            dest_pp_sheet['M19'] = day_net_obj['Prime']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M42'] = day_net_obj['Prime']['Syfy']['target_index']
            dest_pp_sheet['M66'] = day_net_obj['Prime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M91'] = day_net_obj['Prime']['Syfy']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('Telemundo'):
            dest_pp_sheet['N19'] = day_net_obj['Prime']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N42'] = day_net_obj['Prime']['Telemundo']['target_index']
            dest_pp_sheet['N66'] = day_net_obj['Prime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N91'] = day_net_obj['Prime']['Telemundo']['target_frequency']
            pr_im_total += day_net_obj['Prime']['Telemundo']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Prime'].has_key('USA'):
            dest_pp_sheet['O19'] = day_net_obj['Prime']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O42'] = day_net_obj['Prime']['USA']['target_index']
            dest_pp_sheet['O66'] = day_net_obj['Prime']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O91'] = day_net_obj['Prime']['USA']['target_frequency']
            pr_im_total += day_net_obj['Prime']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            pr_re_total += day_net_obj['Prime']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Late Night
        ln_im_total = 0
        ln_re_total = 0
        if day_net_obj['Late Night'].has_key('Bravo'):
            dest_pp_sheet['B20'] = day_net_obj['Late Night']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B43'] = day_net_obj['Late Night']['Bravo']['target_index']
            dest_pp_sheet['B67'] = day_net_obj['Late Night']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B92'] = day_net_obj['Late Night']['Bravo']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Chiller'):
            dest_pp_sheet['C20'] = day_net_obj['Late Night']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C43'] = day_net_obj['Late Night']['Chiller']['target_index']
            dest_pp_sheet['C67'] = day_net_obj['Late Night']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C92'] = day_net_obj['Late Night']['Chiller']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('CNBC'):
            dest_pp_sheet['D20'] = day_net_obj['Late Night']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D43'] = day_net_obj['Late Night']['CNBC']['target_index']
            dest_pp_sheet['D67'] = day_net_obj['Late Night']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D92'] = day_net_obj['Late Night']['CNBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('E!'):
            dest_pp_sheet['E20'] = day_net_obj['Late Night']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E43'] = day_net_obj['Late Night']['E!']['target_index']
            dest_pp_sheet['E67'] = day_net_obj['Late Night']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E92'] = day_net_obj['Late Night']['E!']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Esquire'):
            dest_pp_sheet['F20'] = day_net_obj['Late Night']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F43'] = day_net_obj['Late Night']['Esquire']['target_index']
            dest_pp_sheet['F67'] = day_net_obj['Late Night']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F92'] = day_net_obj['Late Night']['Esquire']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Golf Channel'):
            dest_pp_sheet['G20'] = day_net_obj['Late Night']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G43'] = day_net_obj['Late Night']['Golf Channel']['target_index']
            dest_pp_sheet['G67'] = day_net_obj['Late Night']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G92'] = day_net_obj['Late Night']['Golf Channel']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('MSNBC'):
            dest_pp_sheet['H20'] = day_net_obj['Late Night']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H43'] = day_net_obj['Late Night']['MSNBC']['target_index']
            dest_pp_sheet['H67'] = day_net_obj['Late Night']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H92'] = day_net_obj['Late Night']['MSNBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('NBC'):
            dest_pp_sheet['I20'] = day_net_obj['Late Night']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I43'] = day_net_obj['Late Night']['NBC']['target_index']
            dest_pp_sheet['I67'] = day_net_obj['Late Night']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I92'] = day_net_obj['Late Night']['NBC']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('NBCSN'):
            dest_pp_sheet['J20'] = day_net_obj['Late Night']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J43'] = day_net_obj['Late Night']['NBCSN']['target_index']
            dest_pp_sheet['J67'] = day_net_obj['Late Night']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J92'] = day_net_obj['Late Night']['NBCSN']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            ln_re_total += day_net_obj['Late Night']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Oxygen'):
            dest_pp_sheet['K20'] = day_net_obj['Late Night']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K43'] = day_net_obj['Late Night']['Oxygen']['target_index']
            dest_pp_sheet['K67'] = day_net_obj['Late Night']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K92'] = day_net_obj['Late Night']['Oxygen']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Sprout'):
            dest_pp_sheet['L20'] = day_net_obj['Late Night']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L43'] = day_net_obj['Late Night']['Sprout']['target_index']
            dest_pp_sheet['L67'] = day_net_obj['Late Night']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L92'] = day_net_obj['Late Night']['Sprout']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Syfy'):
            dest_pp_sheet['M20'] = day_net_obj['Late Night']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M43'] = day_net_obj['Late Night']['Syfy']['target_index']
            dest_pp_sheet['M67'] = day_net_obj['Late Night']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M92'] = day_net_obj['Late Night']['Syfy']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('Telemundo'):
            dest_pp_sheet['N20'] = day_net_obj['Late Night']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N43'] = day_net_obj['Late Night']['Telemundo']['target_index']
            dest_pp_sheet['N67'] = day_net_obj['Late Night']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N92'] = day_net_obj['Late Night']['Telemundo']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['Telemundo']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj['Late Night'].has_key('USA'):
            dest_pp_sheet['O20'] = day_net_obj['Late Night']['USA']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['O43'] = day_net_obj['Late Night']['USA']['target_index']
            dest_pp_sheet['O67'] = day_net_obj['Late Night']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O92'] = day_net_obj['Late Night']['USA']['target_frequency']
            ln_im_total += day_net_obj['Late Night']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            ln_re_total += day_net_obj['Late Night']['USA']['target_reach'] / net_obj['Total']['target_reach']

        # Overnight
        on_im_total = 0
        on_re_total = 0
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Bravo'):
            dest_pp_sheet['B21'] = day_net_obj['Overnight']['Bravo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['B44'] = day_net_obj['Overnight']['Bravo']['target_index']
            dest_pp_sheet['B68'] = day_net_obj['Overnight']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['B93'] = day_net_obj['Overnight']['Bravo']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Bravo']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Bravo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Chiller'):
            dest_pp_sheet['C21'] = day_net_obj['Overnight']['Chiller']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['C44'] = day_net_obj['Overnight']['Chiller']['target_index']
            dest_pp_sheet['C68'] = day_net_obj['Overnight']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['C93'] = day_net_obj['Overnight']['Chiller']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Chiller']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Chiller']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('CNBC'):
            dest_pp_sheet['D21'] = day_net_obj['Overnight']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['D44'] = day_net_obj['Overnight']['CNBC']['target_index']
            dest_pp_sheet['D68'] = day_net_obj['Overnight']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['D93'] = day_net_obj['Overnight']['CNBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['CNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['CNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('E!'):
            dest_pp_sheet['E21'] = day_net_obj['Overnight']['E!']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['E44'] = day_net_obj['Overnight']['E!']['target_index']
            dest_pp_sheet['E68'] = day_net_obj['Overnight']['E!']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['E93'] = day_net_obj['Overnight']['E!']['target_frequency']
            on_im_total += day_net_obj['Overnight']['E!']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['E!']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Esquire'):
            dest_pp_sheet['F21'] = day_net_obj['Overnight']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['F44'] = day_net_obj['Overnight']['Esquire']['target_index']
            dest_pp_sheet['F68'] = day_net_obj['Overnight']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['F93'] = day_net_obj['Overnight']['Esquire']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Esquire']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Esquire']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Golf Channel'):
            dest_pp_sheet['G21'] = day_net_obj['Overnight']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['G44'] = day_net_obj['Overnight']['Golf Channel']['target_index']
            dest_pp_sheet['G68'] = day_net_obj['Overnight']['Golf Channel']['target_reach'] / net_obj['Total'][
                'target_reach']
            dest_pp_sheet['G93'] = day_net_obj['Overnight']['Golf Channel']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Golf Channel']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            on_re_total += day_net_obj['Overnight']['Golf Channel']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('MSNBC'):
            dest_pp_sheet['H21'] = day_net_obj['Overnight']['MSNBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['H44'] = day_net_obj['Overnight']['MSNBC']['target_index']
            dest_pp_sheet['H68'] = day_net_obj['Overnight']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['H93'] = day_net_obj['Overnight']['MSNBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['MSNBC']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['MSNBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('NBC'):
            dest_pp_sheet['I21'] = day_net_obj['Overnight']['NBC']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['I44'] = day_net_obj['Overnight']['NBC']['target_index']
            dest_pp_sheet['I68'] = day_net_obj['Overnight']['NBC']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['I93'] = day_net_obj['Overnight']['NBC']['target_frequency']
            on_im_total += day_net_obj['Overnight']['NBC']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['NBC']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('NBCSN'):
            dest_pp_sheet['J21'] = day_net_obj['Overnight']['NBCSN']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['J44'] = day_net_obj['Overnight']['NBCSN']['target_index']
            dest_pp_sheet['J68'] = day_net_obj['Overnight']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['J93'] = day_net_obj['Overnight']['NBCSN']['target_frequency']
            on_im_total += day_net_obj['Overnight']['NBCSN']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['NBCSN']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Oxygen'):
            dest_pp_sheet['K21'] = day_net_obj['Overnight']['Oxygen']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['K44'] = day_net_obj['Overnight']['Oxygen']['target_index']
            dest_pp_sheet['K68'] = day_net_obj['Overnight']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['K93'] = day_net_obj['Overnight']['Oxygen']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Oxygen']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Oxygen']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Sprout'):
            dest_pp_sheet['L21'] = day_net_obj['Overnight']['Sprout']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['L44'] = day_net_obj['Overnight']['Sprout']['target_index']
            dest_pp_sheet['L68'] = day_net_obj['Overnight']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['L93'] = day_net_obj['Overnight']['Sprout']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Sprout']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Sprout']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Syfy'):
            dest_pp_sheet['M21'] = day_net_obj['Overnight']['Syfy']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['M44'] = day_net_obj['Overnight']['Syfy']['target_index']
            dest_pp_sheet['M68'] = day_net_obj['Overnight']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['M93'] = day_net_obj['Overnight']['Syfy']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Syfy']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Syfy']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('Telemundo'):
            dest_pp_sheet['N21'] = day_net_obj['Overnight']['Telemundo']['target_impressions'] / net_obj['Total'][
                'target_impressions']
            dest_pp_sheet['N44'] = day_net_obj['Overnight']['Telemundo']['target_index']
            dest_pp_sheet['N68'] = day_net_obj['Overnight']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['N93'] = day_net_obj['Overnight']['Telemundo']['target_frequency']
            on_im_total += day_net_obj['Overnight']['Telemundo']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['Telemundo']['target_reach'] / net_obj['Total']['target_reach']
        if day_net_obj.has_key('Overnight') and day_net_obj['Overnight'].has_key('USA'):
            dest_pp_sheet['O21'] = day_net_obj['Overnight']['USA']['target_impressions'] / net_obj['Total'][
                'target_impOessions']
            dest_pp_sheet['O44'] = day_net_obj['Overnight']['USA']['target_index']
            dest_pp_sheet['O68'] = day_net_obj['Overnight']['USA']['target_reach'] / net_obj['Total']['target_reach']
            dest_pp_sheet['O93'] = day_net_obj['Overnight']['USA']['target_frequency']
            on_im_total += day_net_obj['Overnight']['USA']['target_impressions'] / net_obj['Total']['target_impressions']
            on_re_total += day_net_obj['Overnight']['USA']['target_reach'] / net_obj['Total']['target_reach']

        dest_pp_sheet['P16'] = mo_im_total
        dest_pp_sheet['P17'] = dy_im_total
        dest_pp_sheet['P18'] = fe_im_total
        dest_pp_sheet['P19'] = pr_im_total
        dest_pp_sheet['P20'] = ln_im_total
        dest_pp_sheet['P21'] = on_im_total

        dest_pp_sheet['P63'] = mo_re_total
        dest_pp_sheet['P64'] = dy_re_total
        dest_pp_sheet['P65'] = fe_re_total
        dest_pp_sheet['P66'] = pr_re_total
        dest_pp_sheet['P67'] = ln_re_total
        dest_pp_sheet['P68'] = on_re_total

        dest_pp_sheet['P39'] = day_net_obj['Morning']['Total']['target_index'] if day_net_obj['Morning'].has_key(
            'Total') else ''
        dest_pp_sheet['P40'] = day_net_obj['Daytime']['Total']['target_index'] if day_net_obj['Daytime'].has_key(
            'Total') else ''
        dest_pp_sheet['P41'] = day_net_obj['Early Fringe']['Total']['target_index'] if day_net_obj['Early Fringe'].has_key(
            'Total') else ''
        dest_pp_sheet['P42'] = day_net_obj['Prime']['Total']['target_index'] if day_net_obj['Prime'].has_key(
            'Total') else ''
        dest_pp_sheet['P43'] = day_net_obj['Late Night']['Total']['target_index'] if day_net_obj['Late Night'].has_key(
            'Total') else ''
        dest_pp_sheet['P44'] = day_net_obj['Overnight']['Total']['target_index'] if day_net_obj['Overnight'].has_key(
            'Total') else ''

        dest_pp_sheet['P88'] = day_net_obj['Morning']['Total']['target_frequency'] if day_net_obj['Morning'].has_key(
            'Total') else ''
        dest_pp_sheet['P89'] = day_net_obj['Daytime']['Total']['target_frequency'] if day_net_obj['Daytime'].has_key(
            'Total') else ''
        dest_pp_sheet['P90'] = day_net_obj['Early Fringe']['Total']['target_frequency'] if day_net_obj[
            'Early Fringe'].has_key('Total') else ''
        dest_pp_sheet['P91'] = day_net_obj['Prime']['Total']['target_frequency'] if day_net_obj['Prime'].has_key(
            'Total') else ''
        dest_pp_sheet['P92'] = day_net_obj['Late Night']['Total']['target_frequency'] if day_net_obj['Late Night'].has_key(
            'Total') else ''
        dest_pp_sheet['P93'] = day_net_obj['Overnight']['Total']['target_frequency'] if day_net_obj['Overnight'].has_key(
            'Total') else ''

        # Totals
        dest_pp_sheet['B22'] = net_obj['Bravo']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C22'] = net_obj['Chiller']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D22'] = net_obj['CNBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E22'] = net_obj['E!']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F22'] = net_obj['Esquire']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G22'] = net_obj['Golf Channel']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H22'] = net_obj['MSNBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I22'] = net_obj['NBC']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J22'] = net_obj['NBCSN']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K22'] = net_obj['Oxygen']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L22'] = net_obj['Sprout']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M22'] = net_obj['Syfy']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N22'] = net_obj['Telemundo']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['O22'] = net_obj['USA']['target_impressions'] / net_obj['Total'][
            'target_impressions'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['P22'] = net_obj['Total']['target_impressions'] / net_obj['Total']['target_impressions']

        dest_pp_sheet['B45'] = net_obj['Bravo']['target_index'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C45'] = net_obj['Chiller']['target_index'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D45'] = net_obj['CNBC']['target_index'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E45'] = net_obj['E!']['target_index'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F45'] = net_obj['Esquire']['target_index'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G45'] = net_obj['Golf Channel']['target_index'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H45'] = net_obj['MSNBC']['target_index'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I45'] = net_obj['NBC']['target_index'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J45'] = net_obj['NBCSN']['target_index'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K45'] = net_obj['Oxygen']['target_index'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L45'] = net_obj['Sprout']['target_index'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M45'] = net_obj['Syfy']['target_index'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N45'] = net_obj['Telemundo']['target_index'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['O45'] = net_obj['USA']['target_index'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['P45'] = net_obj['Total']['target_index']

        dest_pp_sheet['B69'] = net_obj['Bravo']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Bravo') else ''
        dest_pp_sheet['C69'] = net_obj['Chiller']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Chiller') else ''
        dest_pp_sheet['D69'] = net_obj['CNBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'CNBC') else ''
        dest_pp_sheet['E69'] = net_obj['E!']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'E!') else ''
        dest_pp_sheet['F69'] = net_obj['Esquire']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Esquire') else ''
        dest_pp_sheet['G69'] = net_obj['Golf Channel']['target_reach'] / net_obj['Total'][
            'target_reach'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H69'] = net_obj['MSNBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'MSNBC') else ''
        dest_pp_sheet['I69'] = net_obj['NBC']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'NBC') else ''
        dest_pp_sheet['J69'] = net_obj['NBCSN']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'NBCSN') else ''
        dest_pp_sheet['K69'] = net_obj['Oxygen']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Oxygen') else ''
        dest_pp_sheet['L69'] = net_obj['Sprout']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Sprout') else ''
        dest_pp_sheet['M69'] = net_obj['Syfy']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Syfy') else ''
        dest_pp_sheet['N69'] = net_obj['Telemundo']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'Telemundo') else ''
        dest_pp_sheet['O69'] = net_obj['USA']['target_reach'] / net_obj['Total']['target_reach'] if net_obj.has_key(
            'USA') else ''
        dest_pp_sheet['P69'] = net_obj['Total']['target_reach'] / net_obj['Total']['target_reach']

        dest_pp_sheet['B94'] = net_obj['Bravo']['target_frequency'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C94'] = net_obj['Chiller']['target_frequency'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['D94'] = net_obj['CNBC']['target_frequency'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['E94'] = net_obj['E!']['target_frequency'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['F94'] = net_obj['Esquire']['target_frequency'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['G94'] = net_obj['Golf Channel']['target_frequency'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['H94'] = net_obj['MSNBC']['target_frequency'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['I94'] = net_obj['NBC']['target_frequency'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['J94'] = net_obj['NBCSN']['target_frequency'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['K94'] = net_obj['Oxygen']['target_frequency'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['L94'] = net_obj['Sprout']['target_frequency'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['M94'] = net_obj['Syfy']['target_frequency'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['N94'] = net_obj['Telemundo']['target_frequency'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['N94'] = net_obj['USA']['target_frequency'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['O94'] = net_obj['Total']['target_frequency']

        dest_pp_sheet['C100'] = net_obj['Bravo']['target_frequency'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C101'] = net_obj['Chiller']['target_frequency'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['C102'] = net_obj['CNBC']['target_frequency'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['C103'] = net_obj['E!']['target_frequency'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['C104'] = net_obj['Esquire']['target_frequency'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['C105'] = net_obj['Golf Channel']['target_frequency'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['C106'] = net_obj['MSNBC']['target_frequency'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['C107'] = net_obj['NBC']['target_frequency'] if net_obj.has_key('NBC') else ''
        dest_pp_sheet['C108'] = net_obj['NBCSN']['target_frequency'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['C109'] = net_obj['Oxygen']['target_frequency'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['C110'] = net_obj['Sprout']['target_frequency'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['C111'] = net_obj['Syfy']['target_frequency'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['C112'] = net_obj['Telemundo']['target_frequency'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['C113'] = net_obj['USA']['target_frequency'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['C99'] = net_obj['Total']['target_frequency']

        dest_pp_sheet['C117'] = day_net_obj['Morning']['NBC']['target_frequency'] if day_net_obj['Morning'].has_key(
            'NBC') else ''
        dest_pp_sheet['C118'] = day_net_obj['Daytime']['NBC']['target_frequency'] if day_net_obj['Daytime'].has_key(
            'NBC') else ''
        dest_pp_sheet['C119'] = day_net_obj['Early Fringe']['NBC']['target_frequency'] if day_net_obj[
            'Early Fringe'].has_key('NBC') else ''
        dest_pp_sheet['C120'] = day_net_obj['Prime']['NBC']['target_frequency'] if day_net_obj['Prime'].has_key(
            'NBC') else ''
        dest_pp_sheet['C121'] = day_net_obj['Late Night']['NBC']['target_frequency'] if day_net_obj['Late Night'].has_key(
            'NBC') else ''
        dest_pp_sheet['C122'] = day_net_obj['Overnight']['NBC']['target_frequency'] if day_net_obj.has_key('Overnight') and \
                                                                                       day_net_obj['Overnight'].has_key(
                                                                                           'NBC') else ''

        dest_pp_sheet['A126'] = net_obj['Bravo']['tCPM'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['B126'] = net_obj['Bravo']['target_reach'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['C126'] = net_obj['Bravo']['target_frequency'] if net_obj.has_key('Bravo') else ''
        dest_pp_sheet['A127'] = net_obj['Chiller']['tCPM'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['B127'] = net_obj['Chiller']['target_reach'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['C127'] = net_obj['Chiller']['target_frequency'] if net_obj.has_key('Chiller') else ''
        dest_pp_sheet['A128'] = net_obj['CNBC']['tCPM'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['B128'] = net_obj['CNBC']['target_reach'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['C128'] = net_obj['CNBC']['target_frequency'] if net_obj.has_key('CNBC') else ''
        dest_pp_sheet['A129'] = net_obj['E!']['tCPM'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['B129'] = net_obj['E!']['target_reach'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['C129'] = net_obj['E!']['target_frequency'] if net_obj.has_key('E!') else ''
        dest_pp_sheet['A130'] = net_obj['Esquire']['tCPM'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['B130'] = net_obj['Esquire']['target_reach'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['C130'] = net_obj['Esquire']['target_frequency'] if net_obj.has_key('Esquire') else ''
        dest_pp_sheet['A131'] = net_obj['Golf Channel']['tCPM'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['B131'] = net_obj['Golf Channel']['target_reach'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['C131'] = net_obj['Golf Channel']['target_frequency'] if net_obj.has_key('Golf Channel') else ''
        dest_pp_sheet['A132'] = net_obj['MSNBC']['tCPM'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['B132'] = net_obj['MSNBC']['target_reach'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['C132'] = net_obj['MSNBC']['target_frequency'] if net_obj.has_key('MSNBC') else ''
        dest_pp_sheet['A133'] = net_obj['NBCSN']['tCPM'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['B133'] = net_obj['NBCSN']['target_reach'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['C133'] = net_obj['NBCSN']['target_frequency'] if net_obj.has_key('NBCSN') else ''
        dest_pp_sheet['A134'] = net_obj['Oxygen']['tCPM'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['B134'] = net_obj['Oxygen']['target_reach'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['C134'] = net_obj['Oxygen']['target_frequency'] if net_obj.has_key('Oxygen') else ''
        dest_pp_sheet['A135'] = net_obj['Sprout']['tCPM'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['B135'] = net_obj['Sprout']['target_reach'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['C135'] = net_obj['Sprout']['target_frequency'] if net_obj.has_key('Sprout') else ''
        dest_pp_sheet['A136'] = net_obj['Syfy']['tCPM'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['B136'] = net_obj['Syfy']['target_reach'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['C136'] = net_obj['Syfy']['target_frequency'] if net_obj.has_key('Syfy') else ''
        dest_pp_sheet['A137'] = net_obj['Telemundo']['tCPM'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['B137'] = net_obj['Telemundo']['target_reach'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['C137'] = net_obj['Telemundo']['target_frequency'] if net_obj.has_key('Telemundo') else ''
        dest_pp_sheet['A138'] = net_obj['USA']['tCPM'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['B138'] = net_obj['USA']['target_reach'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['C138'] = net_obj['USA']['target_frequency'] if net_obj.has_key('USA') else ''
        dest_pp_sheet['A139'] = day_net_obj['Morning']['NBC']['tCPM'] if day_net_obj['Morning'].has_key('NBC') else ''
        dest_pp_sheet['B139'] = day_net_obj['Morning']['NBC']['target_reach'] if day_net_obj['Morning'].has_key(
            'NBC') else ''
        dest_pp_sheet['C139'] = day_net_obj['Morning']['NBC']['target_frequency'] if day_net_obj['Morning'].has_key(
            'NBC') else ''
        dest_pp_sheet['A140'] = day_net_obj['Daytime']['NBC']['tCPM'] if day_net_obj['Daytime'].has_key('NBC') else ''
        dest_pp_sheet['B140'] = day_net_obj['Daytime']['NBC']['target_reach'] if day_net_obj['Daytime'].has_key(
            'NBC') else ''
        dest_pp_sheet['C140'] = day_net_obj['Daytime']['NBC']['target_frequency'] if day_net_obj['Daytime'].has_key(
            'NBC') else ''
        dest_pp_sheet['A141'] = day_net_obj['Early Fringe']['NBC']['tCPM'] if day_net_obj['Early Fringe'].has_key(
            'NBC') else ''
        dest_pp_sheet['B141'] = day_net_obj['Early Fringe']['NBC']['target_reach'] if day_net_obj['Early Fringe'].has_key(
            'NBC') else ''
        dest_pp_sheet['C141'] = day_net_obj['Early Fringe']['NBC']['target_frequency'] if day_net_obj[
            'Early Fringe'].has_key('NBC') else ''
        dest_pp_sheet['A142'] = day_net_obj['Prime']['NBC']['tCPM'] if day_net_obj['Prime'].has_key('NBC') else ''
        dest_pp_sheet['B142'] = day_net_obj['Prime']['NBC']['target_reach'] if day_net_obj['Prime'].has_key('NBC') else ''
        dest_pp_sheet['C142'] = day_net_obj['Prime']['NBC']['target_frequency'] if day_net_obj['Prime'].has_key(
            'NBC') else ''
        dest_pp_sheet['A143'] = day_net_obj['Late Night']['NBC']['tCPM'] if day_net_obj.has_key('Overnight') and \
                                                                            day_net_obj['Late Night'].has_key('NBC') else ''
        dest_pp_sheet['B143'] = day_net_obj['Late Night']['NBC']['target_reach'] if day_net_obj.has_key('Overnight') and \
                                                                                    day_net_obj['Late Night'].has_key(
                                                                                        'NBC') else ''
        dest_pp_sheet['C143'] = day_net_obj['Late Night']['NBC']['target_frequency'] if day_net_obj.has_key('Overnight') and \
                                                                                        day_net_obj['Late Night'].has_key(
                                                                                            'NBC') else ''
        dest_pp_sheet['A144'] = day_net_obj['Overnight']['NBC']['tCPM'] if day_net_obj.has_key('Overnight') and day_net_obj[
            'Overnight'].has_key('NBC') else ''
        dest_pp_sheet['B144'] = day_net_obj['Overnight']['NBC']['target_reach'] if day_net_obj.has_key('Overnight') and \
                                                                                   day_net_obj['Overnight'].has_key(
                                                                                       'NBC') else ''
        dest_pp_sheet['C144'] = day_net_obj['Overnight']['NBC']['target_frequency'] if day_net_obj.has_key('Overnight') and \
                                                                                       day_net_obj['Overnight'].has_key(
                                                                                           'NBC') else ''

    summary_wb.save(filename)
    return True


def move_when_done(processed, summary_equiv, summary_unequiv):
    shutil.move('./preprocessed/' + processed, './processed/' + processed)
    shutil.move(summary_equiv, './summaries/' + summary_equiv)
    shutil.move(summary_unequiv, './summaries/' + summary_unequiv)
    return True


def merge_opt_and_unopt():
    done_client = []
    listing = glob.glob('./summaries/*.xlsx')
    for filename in listing:
        filename = os.path.basename(filename)
        client = re.split('_', filename)[0]

        if client not in done_client:
            client_listing = glob.glob('./summaries/' + client + '*_unequiv*.xlsx')
            print "merging"
            print client_listing
            copy_unopt_and_opt(client_listing)
            client_listing = glob.glob('./summaries/' + client + '*_equiv*.xlsx')
            print "merging"
            print client_listing
            copy_unopt_and_opt(client_listing)

            done_client.append(client)
            print "done merging " + client


def copy_unopt_and_opt(files_list):
    for sum_file in files_list:
        if 'normalized' in sum_file:
            unopt_file = sum_file
        else:
            opt_file = sum_file

    unopt_wb = load_workbook(unopt_file)
    opt_wb = load_workbook(opt_file)

    unopt_sheet = unopt_wb.get_sheet_by_name('Powerpoint Data')
    opt_sheet = opt_wb.get_sheet_by_name('Powerpoint Data')

    # Copy unopt Powerpoint sheet to opt
    for row_num in range(1, unopt_sheet.max_row + 1):
        opt_sheet['B' + str(row_num)] = unopt_sheet['B' + str(row_num)].value

    # Copy opt Powerpoint sheet to uopt
    for row_num in range(1, opt_sheet.max_row + 1):
        unopt_sheet['C' + str(row_num)] = opt_sheet['C' + str(row_num)].value

    unopt_sheet = unopt_wb.get_sheet_by_name('Appendix')
    opt_sheet = opt_wb.get_sheet_by_name('Appendix')

    # Copy unopt Appendix sheet to opt
    for row_num in range(5, 12):
        for col_num in range(1,unopt_sheet.max_column + 1):
            opt_sheet.cell(row=row_num, column=col_num).value = unopt_sheet.cell(row=row_num, column=col_num).value

    for row_num in range(28, 36):
        for col_num in range(1,unopt_sheet.max_column + 1):
            opt_sheet.cell(row=row_num, column=col_num).value = unopt_sheet.cell(row=row_num, column=col_num).value

    for row_num in range(52, 60):
        for col_num in range(1,unopt_sheet.max_column + 1):
            opt_sheet.cell(row=row_num, column=col_num).value = unopt_sheet.cell(row=row_num,
                                                                       column=col_num).value
    for row_num in range(76, 84):
        for col_num in range(1,unopt_sheet.max_column + 1):
            opt_sheet.cell(row=row_num, column=col_num).value = unopt_sheet.cell(row=row_num,
                                                                       column=col_num).value
    for row_num in range(99, 124):
        opt_sheet['B' + str(row_num)] = unopt_sheet['B' + str(row_num)].value

    # Copy opt Appendix sheet to unopt
    for row_num in range(16, 24):
        for col_num in range(1,opt_sheet.max_column + 1):
            unopt_sheet.cell(row=row_num, column=col_num).value = opt_sheet.cell(row=row_num, column=col_num).value

    for row_num in range(39, 46):
        for col_num in range(1,unopt_sheet.max_column + 1):
            unopt_sheet.cell(row=row_num, column=col_num).value = opt_sheet.cell(row=row_num, column=col_num).value

    for row_num in range(63, 70):
        for col_num in range(1,unopt_sheet.max_column + 1):
            unopt_sheet.cell(row=row_num, column=col_num).value = opt_sheet.cell(row=row_num,
                                                                       column=col_num).value
    for row_num in range(88, 95):
        for col_num in range(1,opt_sheet.max_column + 1):
            unopt_sheet.cell(row=row_num, column=col_num).value = opt_sheet.cell(row=row_num, column=col_num).value

    for row_num in range(99,124):
        unopt_sheet['C' + str(row_num)] = opt_sheet['C' + str(row_num)].value

    for row_num in range(126, 145):
        for col_num in range(1,4):
            unopt_sheet.cell(row=row_num, column=col_num).value = opt_sheet.cell(row=row_num, column=col_num).value

    unopt_wb.save(unopt_file)
    opt_wb.save(opt_file)

listing = glob.glob('./preprocessed/*.xlsx')
for filename in listing:
    filename = os.path.basename(filename)
    if os.path.isfile('./preprocessed/' + filename):
        print "processing unequiv"
        new_filename_unequiv = setup(filename, False)
        if not new_filename_unequiv:
            print "error setting up " + filename
        print process_summary_tab(new_filename_unequiv, False)
        print process_Network_Daypart_tab(new_filename_unequiv, False)
        print process_frequency_distribution_tab(new_filename_unequiv, False)
        print process_reach_by_week_tab(new_filename_unequiv, False)
        print process_frequency_distribution_by_net_tab(new_filename_unequiv, False)
        print process_network_reach_tab(new_filename_unequiv, False)
        print process_powerpoint_tab(new_filename_unequiv, False)
        print process_appendix_tab(new_filename_unequiv, False)
        print "processing equiv"
        new_filename_equiv = setup(filename, True)
        if not new_filename_equiv:
            print "error setting up " + filename
        print process_summary_tab(new_filename_equiv, True)
        print process_Network_Daypart_tab(new_filename_equiv, True)
        print process_frequency_distribution_tab(new_filename_equiv, True)
        print process_reach_by_week_tab(new_filename_equiv, True)
        print process_frequency_distribution_by_net_tab(new_filename_equiv, True)
        print process_network_reach_tab(new_filename_equiv, True)
        print process_powerpoint_tab(new_filename_equiv, True)
        print process_appendix_tab(new_filename_equiv, True)
        print move_when_done(filename, new_filename_equiv, new_filename_unequiv)

merge_opt_and_unopt()
